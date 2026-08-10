#!/usr/bin/env python3
"""키워드 블랙리스트 판정 1벌 — `제외키워드`(정확일치) · `제외카테고리`(부분일치).

**왜 eroomlib 인가**: 소비자가 스킬 경계를 넘는다.
  - `product-name/scripts/cat_view.py`  → 제외키워드로 뷰 행을 버린다(상표 리스크 차단)
  - `bulsaja-category-fix/scripts/category_gate.py` → 제외카테고리로 상품을 뺀다
  - 나중에 붙을 catfix 게이트 2곳(`bulsaja_mcp.py` · `category_steer.py`)
스킬 안에 두면 옆 스킬이 상대 스크립트를 import 하게 된다.

**매칭 규칙은 `sellerlife-keyword/scripts/detect.py:265-287` 과 글자 그대로 같다.**
  - 제외카테고리: `조각 in 카테고리경로` — **정규화하지 않는다**
  - 제외키워드:   정확일치 (앞뒤 공백만 제거)

정규화를 얹으면 안 되는 이유가 두 개다.
  ① 실측(2026-08-08)에서 조각 207개의 `>` 표기가 불사자 무공백 `>` 형식
     (`생활/건강>공구>운반용품>핸드카트/운반기`)과 그대로 맞물렸다. "조용한 0건"이 아니다.
     여기서만 정규화하면 그 실측(3,161건 중 891건 = 28.2%)이 무효가 된다.
  ② "블랙리스트에 조각을 넣으면 키워드 소싱과 상품 게이트에 **같이** 걸린다"가 이 파일의
     계약이다. 한쪽만 규칙이 다르면 그 계약이 조용히 깨진다.

**AI 판정은 여기 없다.** Gemini 상표 판정은 `detect.py` 안에서 통다운 가공 때만 돈다.
그 결과를 이룸님이 승인하면 `blacklist_update.py` 가 시트에 append 한다.
이 모듈은 **그 명단을 조회만** 한다.

정본 = 드라이브 `50-소스자료/52-키워드-블랙리스트/keyword_blacklist.xlsx` 한 곳(2026-08-08).

CLI:
    python -m eroomlib.exclusion stats
    python -m eroomlib.exclusion check "생활/건강>공구>안전용품>기타안전용품"
    python -m eroomlib.exclusion check-kw 다이슨
"""
import os
import sys

try:  # 패키지로 import 될 때
    from .gdrive import blacklist_cache_path, ensure_blacklist
except ImportError:  # 파일을 직접 실행할 때
    sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
    from eroomlib.gdrive import blacklist_cache_path, ensure_blacklist

SHEET_KW = "제외키워드"
SHEET_CAT = "제외카테고리"
SHEET_BRAND = "제외브랜드"

# 경로를 손으로 지정하는 탈출구(오프라인·재현 검증용). 인자 > 환경변수 > 드라이브 순.
ENV_PATH = "EROOM_BLACKLIST"

# 프로세스 캐시 {경로: 로드결과}.
# **없으면 안 된다**: `ensure_blacklist()` 는 호출마다 드라이브 md5 를 조회하는데,
# 뷰 생성은 카테고리마다 부른다(54그룹 × N카테고리) → 네트워크 왕복이 수백 회가 된다.
# 실측: 로드 84ms(실행당 1회) / 300행 필터 0.005ms. 캐시할 값은 로드 쪽뿐이다.
_cache = {}


def blacklist_path(path=None):
    """블랙리스트 파일 경로. 정본은 드라이브 한 곳, 못 읽으면 **직전 캐시**로 폴백한다.

    폴백이 로컬 `DATA_ROOT/keyword_blacklist/` 가 아니라 드라이브 캐시인 이유:
    그 로컬 경로는 sellerlife-keyword 스킬의 통다운 사정(그 스킬 전용 DATA_ROOT)이고,
    2026-08-08 정본 단일화 때 백업본만 남긴 자리다. 여기서 그걸 다시 읽으면 정본을
    한 곳으로 모은 방향을 되돌리게 된다. 캐시는 **같은 파일의 옛 리비전**이라 안전하다.
    """
    p = path or os.environ.get(ENV_PATH)
    if p:
        if not os.path.exists(p):
            raise FileNotFoundError(f"블랙리스트를 찾지 못했습니다: {p}")
        return p
    try:
        return ensure_blacklist()
    except Exception as e:  # noqa: BLE001 — 네트워크·인증 실패
        cached = blacklist_cache_path()
        if os.path.exists(cached):
            print(f"경고: 드라이브 정본을 못 읽어 직전 캐시를 씁니다 ({str(e)[:80]}). "
                  f"이 실행의 제외 결과는 정본과 다를 수 있습니다 -> {cached}",
                  file=sys.stderr)
            return cached
        raise


def load_blacklist(path=None, refresh=False):
    """블랙리스트 → {"제외키워드": set, "제외카테고리": [str], "제외브랜드": [str], "경로": str}.

    시트 규격은 `detect.py load_blacklist` 와 같다 —
    제외키워드는 A열·1행 헤더, 제외카테고리는 A열·**헤더 없음**, 제외브랜드는 A열·1행 헤더.
    (제외브랜드는 지금 아무도 안 쓴다. 부분일치인데 손실 검증이 없어서다 — 원안 §결정 A.
     읽어만 두는 이유는 통계·리포트에서 "왜 안 쓰나"를 눈으로 확인할 수 있게 하려는 것.)
    """
    p = blacklist_path(path)
    if not refresh and p in _cache:
        return _cache[p]

    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover
        raise ImportError("openpyxl 이 필요합니다: pip install openpyxl>=3.1.0")

    wb = load_workbook(p, read_only=True)
    try:
        kw, cats, brands = set(), [], []

        if SHEET_KW in wb.sheetnames:
            for row in wb[SHEET_KW].iter_rows(min_row=2, min_col=1, max_col=1,
                                              values_only=True):
                v = row[0]
                if v is not None and str(v).strip():
                    kw.add(str(v).strip())
        else:
            print(f"경고: '{SHEET_KW}' 시트가 없습니다 — 키워드 컷 건너뜀", file=sys.stderr)

        if SHEET_CAT in wb.sheetnames:
            for row in wb[SHEET_CAT].iter_rows(min_col=1, max_col=1, values_only=True):
                v = row[0]
                if v is not None and str(v).strip():
                    cats.append(str(v).strip())
        else:
            print(f"경고: '{SHEET_CAT}' 시트가 없습니다 — 카테고리 게이트 건너뜀",
                  file=sys.stderr)

        if SHEET_BRAND in wb.sheetnames:
            for row in wb[SHEET_BRAND].iter_rows(min_row=2, min_col=1, max_col=1,
                                                 values_only=True):
                v = row[0]
                if v is not None and str(v).strip():
                    brands.append(str(v).strip().lower())
    finally:
        wb.close()

    out = {SHEET_KW: kw, SHEET_CAT: cats, SHEET_BRAND: brands, "경로": p}
    _cache[p] = out
    return out


def excluded_keywords(path=None, refresh=False):
    """제외키워드 set — 정확일치용. `kw.strip() in s` 로 쓴다."""
    return load_blacklist(path, refresh)[SHEET_KW]


class CategoryGate:
    """카테고리 경로 부분일치 판정기. 조각 목록을 한 번 받아 여러 번 판정한다."""

    def __init__(self, fragments):
        seen = set()
        uniq = []
        for f in fragments:
            f = str(f).strip()
            if f and f not in seen:
                seen.add(f)
                uniq.append(f)
        # 조각 자체에는 고정 순위를 매기지 않는다 — 어느 쪽이 더 구체적인지는
        # **어느 경로에 걸렸느냐**로만 정해진다(`hits()` 참조). 여기서는 결정성만 준다.
        self.fragments = sorted(uniq)

    def __len__(self):
        return len(self.fragments)

    def hits(self, cat_path):
        """걸린 조각 **전부**, 구체적인 것부터.

        첫 조각만 돌려주지 않는 이유: 어떤 조각이 과매칭인지 세어야 블랙리스트를 고칠
        근거가 생긴다. `제외카테고리` 에는 `여성의류`·`출산/육아` 같은 **대분류급 조각**이
        섞여 있어(369개 중 162개) 그게 혼자 대상을 만드는 상황을 관측해야 한다.

        **"구체적"의 기준 = 그 조각이 경로의 몇 번째 차수에서 걸렸나** (2026-08-09 실측 정정).
        조각 문자열의 생김새(길이·`>` 개수)로는 못 정한다 —

            생활/건강>욕실용품>세면대/수전용품>싱크대수전
              `욕실용품>`      = 2차수에서 걸림, `>` 1개, 5자
              `세면대/수전용품` = 3차수에서 걸림, `>` 0개, 8자   ← 이쪽이 구체적
            생활/건강>공구>펌프>수중펌프
              `생활/건강` = 1차수, 5자
              `>펌프`     = 3차수, 3자                          ← 이쪽이 구체적

        길이순이면 첫 예는 맞고 둘째 예는 틀린다. `>` 개수순이면 반대다. 실측에서 두 기준이
        엇갈린 상품이 1,340건이었고 어느 쪽도 일관되게 맞지 않았다.
        **매치 시작 위치 앞의 `>` 개수**(= 걸린 차수)로 재면 둘 다 맞는다.

        같은 조각이 여러 번 나오면 **첫 매치**를 쓴다. 가장 깊은 매치를 쓰면
        `스포츠/레저>오토바이/스쿠터>오토바이부품>…` 에서 `>오토바이` 가 뒤쪽 `>오토바이부품`
        으로 3차수 판정을 받아 실제 노드명인 `오토바이/스쿠터` 를 이겨 버린다.

        판정(걸리나 안 걸리나)에는 순서가 영향을 주지 않는다 — **대상 집합은 그대로**이고
        기록에 남는 대표 조각만 달라진다.
        """
        s = str(cat_path or "").strip()
        if not s:
            return []
        scored = []
        for f in self.fragments:
            i = s.find(f)
            if i >= 0:
                scored.append((s.count(">", 0, i), len(f), f))
        scored.sort(key=lambda x: (-x[0], -x[1], x[2]))
        return [f for _, _, f in scored]

    def is_excluded(self, cat_path):
        """걸리면 첫(가장 구체적인) 조각, 아니면 None."""
        h = self.hits(cat_path)
        return h[0] if h else None


def category_gate(path=None, refresh=False):
    """블랙리스트 `제외카테고리` 로 만든 게이트."""
    return CategoryGate(load_blacklist(path, refresh)[SHEET_CAT])


def is_broad_fragment(fragment):
    """`>` 가 없는 조각 = 대분류급(`여성의류`·`출산/육아`). 리포트에서 따로 표기한다.

    이런 조각은 카테고리 트리 어느 깊이에서든 걸리므로 대상 수를 혼자 크게 만든다.
    막지는 않는다(이룸님이 넣은 것) — **보이게만** 한다.
    """
    return ">" not in str(fragment or "")


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main():
    import argparse
    ap = argparse.ArgumentParser(description="블랙리스트 판정(제외키워드·제외카테고리)")
    ap.add_argument("cmd", choices=["stats", "check", "check-kw"])
    ap.add_argument("value", nargs="?", help="check=카테고리경로 / check-kw=키워드")
    ap.add_argument("--blacklist", help="블랙리스트 xlsx 경로(기본: 드라이브 정본)")
    args = ap.parse_args()

    bl = load_blacklist(args.blacklist)
    if args.cmd == "stats":
        cats = bl[SHEET_CAT]
        broad = [c for c in cats if is_broad_fragment(c)]
        print(f"경로: {bl['경로']}")
        print(f"제외키워드 {len(bl[SHEET_KW])}개(정확일치) · "
              f"제외카테고리 {len(cats)}개(부분일치, 고유 {len(set(cats))}) · "
              f"제외브랜드 {len(bl[SHEET_BRAND])}개(미사용)")
        print(f"  대분류급 조각('>' 없음) {len(broad)}개: "
              + ", ".join(broad[:15]) + (" …" if len(broad) > 15 else ""))
        return

    if not args.value:
        ap.error(f"{args.cmd} 는 값이 필요합니다")

    if args.cmd == "check-kw":
        hit = args.value.strip() in bl[SHEET_KW]
        print(f"{'제외' if hit else '통과'}  {args.value}")
        sys.exit(1 if hit else 0)

    gate = CategoryGate(bl[SHEET_CAT])
    hits = gate.hits(args.value)
    print(f"{'제외' if hits else '통과'}  {args.value}")
    if hits:
        print("  걸린 조각: " + " · ".join(hits))
    sys.exit(1 if hits else 0)


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    main()
