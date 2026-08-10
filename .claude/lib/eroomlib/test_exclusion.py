#!/usr/bin/env python3
"""블랙리스트 판정 회귀 테스트 — 임시 xlsx 를 만들어 네트워크 없이 돈다.

    python .claude/lib/eroomlib/test_exclusion.py

지키려는 계약:
  1. 카테고리는 **부분일치·정규화 없음** — `detect.py:266` 과 같은 결과를 낸다
  2. 불사자 무공백 `>` 형식에 조각이 그대로 맞물린다 ("조용한 0건"이 아니다)
  3. `hits()` 가 걸린 조각을 **전부** 돌려주고, 첫 조각이 **가장 구체적인 것**이다
  4. 키워드는 **정확일치** — 부분일치로 새지 않는다
  5. 로드는 **프로세스 캐시 1회** (뷰 생성이 카테고리마다 부르므로)
"""
import os
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from eroomlib import exclusion  # noqa: E402

# 실제 블랙리스트에서 가져온 표본 — 대분류급 조각과 세부 조각이 섞여 있다.
CATS = ["생활/건강>공구>안전용품>기타안전용품", "여성의류", ">펌프", "생활/건강", "식품>"]
KWS = ["다이슨", "무선청소기"]


def _make_xlsx(path, cats=CATS, kws=KWS, brands=("dyson",),
               skip=()):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    if exclusion.SHEET_KW not in skip:
        ws = wb.create_sheet(exclusion.SHEET_KW)
        ws.append(["키워드", "분류", "제외사유"])       # 1행 헤더
        for k in kws:
            ws.append([k, "", "상표"])
    if exclusion.SHEET_CAT not in skip:
        ws = wb.create_sheet(exclusion.SHEET_CAT)      # 헤더 없음 — 1행부터 데이터
        for c in cats:
            ws.append([c])
    if exclusion.SHEET_BRAND not in skip:
        ws = wb.create_sheet(exclusion.SHEET_BRAND)
        ws.append(["브랜드어근"])
        for b in brands:
            ws.append([b])
    wb.save(path)
    return path


class Base(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.mkdtemp()
        self.path = _make_xlsx(os.path.join(self.tmp, "keyword_blacklist.xlsx"))
        exclusion._cache.clear()

    def tearDown(self):
        exclusion._cache.clear()


class CategoryTest(Base):
    def test_무공백_부분일치가_붙는다(self):
        """실측 근거: 조각 207개가 `>` 를 포함하는데 불사자도 무공백 `>` 다."""
        g = exclusion.category_gate(self.path)
        self.assertEqual(
            g.is_excluded("생활/건강>공구>안전용품>기타안전용품"),
            "생활/건강>공구>안전용품>기타안전용품")
        self.assertEqual(g.is_excluded("생활/건강>공구>펌프>수중펌프"), ">펌프")

    def test_정규화하지_않는다(self):
        """공백이 낀 표기는 **일부러** 안 걸린다 — detect.py 와 같은 규칙이어야 한다."""
        g = exclusion.category_gate(self.path)
        self.assertIsNone(g.is_excluded("생활 / 건강 > 공구 > 펌프"))

    def test_안_걸리는_경로는_None(self):
        g = exclusion.category_gate(self.path)
        self.assertIsNone(g.is_excluded("디지털/가전>주변기기>마우스"))
        self.assertIsNone(g.is_excluded(""))
        self.assertIsNone(g.is_excluded(None))

    def test_hits는_전부_돌려주고_구체적인_것이_먼저다(self):
        g = exclusion.category_gate(self.path)
        h = g.hits("생활/건강>공구>안전용품>기타안전용품")
        self.assertEqual(h[0], "생활/건강>공구>안전용품>기타안전용품")
        self.assertIn("생활/건강", h)          # 대분류급 조각도 같이 잡힌다
        self.assertEqual(len(h), 2)

    def test_깊은_차수에서_걸린_조각이_먼저다(self):
        """길이로 정렬하면 `생활/건강`(5자)이 `>펌프`(3자)를 이겨 원인 추적이 막힌다."""
        g = exclusion.category_gate(self.path)
        self.assertEqual(g.hits("생활/건강>공구>펌프>수중펌프"), [">펌프", "생활/건강"])

    def test_조각_생김새가_아니라_걸린_차수로_고른다(self):
        """실측 정정(2026-08-09) — 길이순·`>`개수순 어느 쪽도 일관되게 맞지 않았다.

        `욕실용품>`(2차수, `>` 1개)  vs  `세면대/수전용품`(3차수, `>` 0개)
        → `>` 개수순이면 앞엣것이 이기지만, 실제로 더 구체적인 건 뒤엣것이다.
        """
        g = exclusion.CategoryGate(["욕실용품>", "세면대/수전용품"])
        self.assertEqual(g.hits("생활/건강>욕실용품>세면대/수전용품>싱크대수전"),
                         ["세면대/수전용품", "욕실용품>"])
        # 같은 규칙이 반대 모양(`>`가 많은 쪽이 더 깊은 경우)도 맞춘다
        g2 = exclusion.CategoryGate(["생활/건강", ">펌프"])
        self.assertEqual(g2.hits("생활/건강>공구>펌프>수중펌프"), [">펌프", "생활/건강"])

    def test_같은_조각이_여러번이면_첫_매치로_잰다(self):
        """가장 깊은 매치를 쓰면 `>오토바이` 가 실제 노드명 `오토바이/스쿠터` 를 이긴다."""
        g = exclusion.CategoryGate([">오토바이", "오토바이/스쿠터"])
        self.assertEqual(g.hits("스포츠/레저>오토바이/스쿠터>오토바이부품>기타오토바이부품"),
                         ["오토바이/스쿠터", ">오토바이"])

    def test_순서가_바뀌어도_대상_집합은_같다(self):
        """정렬은 기록용이다 — 걸리나 안 걸리나에는 영향이 없다."""
        g = exclusion.category_gate(self.path)
        for cat in ["생활/건강>공구>펌프>수중펌프", "디지털/가전>주변기기>마우스", "여성의류>원피스"]:
            self.assertEqual(bool(g.is_excluded(cat)), bool(g.hits(cat)))
            self.assertEqual(set(g.hits(cat)),
                             {f for f in g.fragments if f in cat})

    def test_대분류급_조각_표시(self):
        self.assertTrue(exclusion.is_broad_fragment("여성의류"))
        self.assertFalse(exclusion.is_broad_fragment("식품>"))

    def test_중복조각은_한_번만(self):
        g = exclusion.CategoryGate(["여성의류", "여성의류", " 여성의류 ", ""])
        self.assertEqual(len(g), 1)


class KeywordTest(Base):
    def test_정확일치만(self):
        s = exclusion.excluded_keywords(self.path)
        self.assertIn("다이슨", s)
        self.assertNotIn("다이슨청소기", s)     # 부분일치로 새면 안 된다
        self.assertNotIn("다이", s)

    def test_헤더행은_안_들어간다(self):
        s = exclusion.excluded_keywords(self.path)
        self.assertNotIn("키워드", s)


class LoadTest(Base):
    def test_프로세스_캐시는_1회만_읽는다(self):
        """ensure_blacklist 는 호출마다 드라이브 md5 를 본다 — 캐시가 없으면 왕복이 터진다."""
        calls = []
        real = exclusion.load_blacklist

        exclusion.load_blacklist(self.path)
        self.assertIn(self.path, exclusion._cache)

        import openpyxl
        orig = openpyxl.load_workbook
        openpyxl.load_workbook = lambda *a, **k: calls.append(1) or orig(*a, **k)
        try:
            for _ in range(5):
                real(self.path)
        finally:
            openpyxl.load_workbook = orig
        self.assertEqual(calls, [], "캐시 이후에는 파일을 다시 열지 않아야 한다")

    def test_refresh는_다시_읽는다(self):
        exclusion.load_blacklist(self.path)
        _make_xlsx(self.path, cats=["새조각>"], kws=["새키워드"])
        bl = exclusion.load_blacklist(self.path, refresh=True)
        self.assertEqual(bl[exclusion.SHEET_CAT], ["새조각>"])

    def test_시트가_없으면_비고_죽지_않는다(self):
        p = _make_xlsx(os.path.join(self.tmp, "no_cat.xlsx"),
                       skip=(exclusion.SHEET_CAT,))
        bl = exclusion.load_blacklist(p)
        self.assertEqual(bl[exclusion.SHEET_CAT], [])
        self.assertTrue(bl[exclusion.SHEET_KW])

    def test_없는_경로는_바로_실패한다(self):
        with self.assertRaises(FileNotFoundError):
            exclusion.load_blacklist(os.path.join(self.tmp, "없음.xlsx"))

    def test_환경변수_경로(self):
        os.environ[exclusion.ENV_PATH] = self.path
        try:
            self.assertEqual(exclusion.blacklist_path(), self.path)
        finally:
            del os.environ[exclusion.ENV_PATH]


class DetectParityTest(Base):
    """`detect.py:265-287` 의 1·2단계와 **같은 판정**을 내는지 직접 대조한다."""

    def test_detect와_동일(self):
        bl = exclusion.load_blacklist(self.path)
        cat_list, kw_set = bl[exclusion.SHEET_CAT], bl[exclusion.SHEET_KW]
        gate = exclusion.CategoryGate(cat_list)
        samples = ["생활/건강>공구>펌프>수중펌프", "여성의류>원피스", "디지털/가전>TV",
                   "식품>가공식품", "생활 / 건강"]
        for cat in samples:
            detect_says = any(frag in cat for frag in cat_list)   # detect.py:266 그대로
            self.assertEqual(bool(gate.is_excluded(cat)), detect_says, cat)
        for kw in ["다이슨", "다이슨청소기", "무선청소기", "청소기"]:
            self.assertEqual(kw in kw_set, kw in KWS, kw)         # detect.py:273 그대로


if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    unittest.main(verbosity=2)
