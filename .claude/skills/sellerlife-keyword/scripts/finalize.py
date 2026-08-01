#!/usr/bin/env python3
"""AI 판정 결과를 최종 산출물로 분리.

detected/*_AI검출.xlsx 를 카테고리별로 묶어 (20만행 상한 때문에 쪼개진
'생활&건강' + '생활&건강2' 같은 분할 파일을 하나로 병합) 두 파일을 만든다.

  최종_<카테고리>.xlsx  = 판정 '안전' 행만, AI 열 5개 제거 -> 원래 13열. 바로 사용 가능.
  검토_<카테고리>.xlsx  = 판정 'AI위험' 행 + AI분류/AI사유 + '승인' 열. 주황 표시.
  최종_통합_<YYMMDD>.xlsx = 최종 3개를 하나로 병합(맨 앞 '소싱카테고리' 열). --no-merge 로 끔.

검토 '승인' 열은 규칙대로 미리 채워진다(--no-prefill 로 끔):
  · '어린이용' 분류 → 키워드에 아동 표현(어린이/유아/키즈…)이 있을 때만 'O', 없으면 빈칸(오탐 살림)
  · 그 외 분류(브랜드/캐릭터/상표/모델 등) → 'O'
사람은 검토 파일을 열어 오탐만 빈칸으로 지우면 된다. blacklist_update.py 는 'O' 만 반영.

사용법:
  python finalize.py --in-dir <detected폴더> --out-dir <run폴더> [--no-prefill] [--no-merge]
"""
import argparse
import glob
import json
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from common import cat_key, is_kid_keyword, norm  # noqa: E402

AI_COLS = ["상태", "AI위험판정", "AI분류", "AI사유", "판정"]
REVIEW_HEADER = ["키워드", "카테고리", "AI분류", "AI사유", "승인"]

# cat_key -> 보기 좋은 표시명
DISPLAY = {"가구인테리어": "가구인테리어", "디지털가전": "디지털가전", "생활건강": "생활건강"}


def _category_of(path):
    """filtered_SellarSourcing_생활&건강2_AI검출.xlsx -> '생활건강'"""
    name = os.path.basename(path)
    for suffix in ("_AI검출.xlsx", ".xlsx"):
        if name.endswith(suffix):
            name = name[: -len(suffix)]
            break
    for prefix in ("filtered_", "SellarSourcing_"):
        if name.startswith(prefix):
            name = name[len(prefix):]
    name = name.replace("SellarSourcing_", "")
    return cat_key(name)


def read_sheet(path):
    """(header, rows) 로 읽는다."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it if r is not None and any(v is not None for v in r)]
    wb.close()
    return header, rows


def col_index(header, name):
    n = norm(name)
    for i, h in enumerate(header):
        if norm(h) == n:
            return i
    return None


def prefill_approve(kw, ai_type):
    """검토 '승인' 열 기본값 제안.

    '어린이용' 분류는 아동 표현이 명시된 것만 'O'(성인 겸용 오탐은 빈칸으로 살림),
    그 외 분류(브랜드/캐릭터/상표/모델 등)는 'O'. 사람이 오탐만 지운다.
    """
    if str(ai_type or "").strip() == "어린이용":
        return "O" if is_kid_keyword(kw) else None
    return "O"


def finalize(in_dir, out_dir, prefill=True, merge=True):
    files = sorted(glob.glob(os.path.join(in_dir, "*_AI검출.xlsx")))
    if not files:
        print(f"[오류] '{in_dir}' 안에 *_AI검출.xlsx 가 없습니다.")
        return None

    groups = {}  # cat_key -> {"header":…, "safe":[…], "risk":[…], "err":n}
    for fp in files:
        ck = _category_of(fp)
        header, rows = read_sheet(fp)

        i_judge = col_index(header, "판정")
        i_type = col_index(header, "AI분류")
        i_reason = col_index(header, "AI사유")
        i_kw = col_index(header, "키워드")
        i_cat = col_index(header, "카테고리")
        if i_judge is None:
            print(f"  [건너뜀] {os.path.basename(fp)}: '판정' 열 없음")
            continue

        ai_idx = {col_index(header, c) for c in AI_COLS} - {None}
        keep_idx = [i for i in range(len(header)) if i not in ai_idx]

        g = groups.setdefault(ck, {"header": [header[i] for i in keep_idx],
                                   "safe": [], "risk": [], "err": 0, "files": []})
        g["files"].append(os.path.basename(fp))

        for row in rows:
            def at(i):
                return row[i] if i is not None and i < len(row) else None

            judge = at(i_judge)
            if judge == "안전":
                g["safe"].append([at(i) for i in keep_idx])
            elif judge == "AI위험":
                g["risk"].append([at(i_kw), at(i_cat), at(i_type), at(i_reason), None])
            elif judge == "AI오류-재실행필요":
                g["err"] += 1

    os.makedirs(out_dir, exist_ok=True)
    fill_risk = PatternFill("solid", fgColor="FFC000")
    summary = []
    merged_header, merged_rows = None, []

    for ck, g in sorted(groups.items()):
        disp = DISPLAY.get(ck, ck)

        # ---- 최종 (안전만, 키워드 중복 제거)
        seen, safe_rows = set(), []
        i_kw = col_index(g["header"], "키워드")
        for row in g["safe"]:
            k = row[i_kw] if i_kw is not None else None
            if k in seen:
                continue
            seen.add(k)
            safe_rows.append(row)

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(g["header"])
        for c in range(1, len(g["header"]) + 1):
            ws.cell(1, c).font = Font(bold=True)
        for row in safe_rows:
            ws.append(row)
        for j, h in enumerate(g["header"], start=1):
            if "해외배송비율" in norm(h):
                letter = get_column_letter(j)
                for r in range(2, len(safe_rows) + 2):
                    ws[f"{letter}{r}"].number_format = "0.00%"
        final_path = os.path.join(out_dir, f"최종_{disp}.xlsx")
        wb.save(final_path)

        # ---- 통합본용 행 수집 (맨 앞에 소싱카테고리)
        if merge:
            if merged_header is None:
                merged_header = ["소싱카테고리"] + list(g["header"])
            for row in safe_rows:
                merged_rows.append([disp] + list(row))

        # ---- 검토 (AI위험, 승인 열 규칙대로 pre-fill)
        seen_r, risk_rows = set(), []
        for row in g["risk"]:
            if row[0] in seen_r:
                continue
            seen_r.add(row[0])
            risk_rows.append(row)

        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Sheet1"
        ws2.append(REVIEW_HEADER)
        for c in range(1, len(REVIEW_HEADER) + 1):
            ws2.cell(1, c).font = Font(bold=True)
        for row in risk_rows:
            if prefill:
                row[4] = prefill_approve(row[0], row[2])
            ws2.append(row)
            ws2.cell(ws2.max_row, 1).fill = fill_risk
        for col, w in zip("ABCDE", [24, 40, 12, 34, 8]):
            ws2.column_dimensions[col].width = w
        ws2.freeze_panes = "A2"
        review_path = os.path.join(out_dir, f"검토_{disp}.xlsx")
        wb2.save(review_path)

        print(f"[{disp}] 안전 {len(safe_rows):,} / AI위험 {len(risk_rows):,} / 오류 {g['err']:,}"
              f"  (원본 {len(g['files'])}개 병합)")
        if g["err"]:
            print(f"    ※ 오류 {g['err']}건 — detect.py 를 다시 돌리면 그 행만 재처리됩니다.")

        summary.append({
            "카테고리": disp, "안전": len(safe_rows), "AI위험": len(risk_rows),
            "오류": g["err"], "최종": final_path, "검토": review_path,
        })

    # ---- 최종 통합본 (최종 3개 -> 하나, 맨 앞 소싱카테고리 열)
    if merge and merged_rows:
        mwb = Workbook()
        mws = mwb.active
        mws.title = "전체"
        mws.append(merged_header)
        for c in range(1, len(merged_header) + 1):
            mws.cell(1, c).font = Font(bold=True)
        for row in merged_rows:
            mws.append(row)
        for j, h in enumerate(merged_header, start=1):
            if "해외배송비율" in norm(h):
                letter = get_column_letter(j)
                for r in range(2, len(merged_rows) + 2):
                    mws[f"{letter}{r}"].number_format = "0.00%"
        mws.freeze_panes = "A2"
        ymd = os.path.basename(os.path.normpath(out_dir))
        tag = f"_{ymd}" if ymd.isdigit() else ""
        merged_path = os.path.join(out_dir, f"최종_통합{tag}.xlsx")
        mwb.save(merged_path)
        print(f"\n[통합] 최종 {len(merged_rows):,}행 -> {os.path.basename(merged_path)}")

    return summary


def main():
    ap = argparse.ArgumentParser(description="최종/검토 엑셀 생성")
    ap.add_argument("--in-dir", required=True, help="detected 폴더 (*_AI검출.xlsx)")
    ap.add_argument("--out-dir", required=True, help="산출물 저장 폴더")
    ap.add_argument("--no-prefill", action="store_true", help="검토 '승인' 열 자동채움 끄기(전부 빈칸)")
    ap.add_argument("--no-merge", action="store_true", help="최종 통합본 생성 끄기")
    args = ap.parse_args()

    summary = finalize(args.in_dir, args.out_dir,
                       prefill=not args.no_prefill, merge=not args.no_merge)
    if not summary:
        sys.exit(1)

    tot_s = sum(s["안전"] for s in summary)
    tot_r = sum(s["AI위험"] for s in summary)
    tot_e = sum(s["오류"] for s in summary)
    print(f"\n합계: 안전 {tot_s:,} / AI위험 {tot_r:,} / 오류 {tot_e:,}")
    print(f"저장 위치: {args.out_dir}")
    print("\n###SUMMARY###")
    print(json.dumps(summary, ensure_ascii=False))


if __name__ == "__main__":
    main()
