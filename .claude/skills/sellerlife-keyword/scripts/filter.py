#!/usr/bin/env python3
"""셀러라이프 rawdata 엑셀 필터.

sellerlife_filter.py 포팅:
  - 원본 54번 줄의 닫히지 않은 따옴표(SyntaxError) 수정
  - pandas.read_excel -> openpyxl read_only 스트리밍 (원본이 최대 260MB, 20만행)
  - 임계값을 CLI 인자로 노출

동작:
  기준1  불필요한 열 22개 삭제 (이름 기준, 줄바꿈/공백 무시하고 매칭) -> 13열
  기준2  브랜드키워드 == 'O'  행 삭제
  기준3  쇼핑성키워드 == 'X'  행 삭제
  기준4  네이버상품수 <= 30000        (숫자 아님 -> 삭제)
  기준5  네이버해외배송비율 >= 0.1    (숫자 아님 -> 삭제)
  기준6  쿠팡해외배송비율 >= 0.1      (숫자 아님 -> 삭제)
  기준7  네이버평균가 <= 5,000,000    (빈칸/0 은 '가격정보 없음' 이므로 유지)

출력 시트명은 반드시 'Sheet1' — detect.py 가 읽는 계약이다.

사용법:
  python filter.py --in-dir <raw폴더> --out-dir <filtered폴더>
  python filter.py --in-dir ... --ship-rate 0.5 --max-price 500000
"""
import argparse
import glob
import os
import sys

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from common import norm, find_col  # noqa: E402

# 삭제할 열 22개 (35열 -> 13열)
COLS_TO_DELETE = [
    "예상1개월검색량",
    "예상1개월검색량상승률",
    "최근3개월검색량",
    "예상3개월검색량",
    "예상3개월검색량상승률",
    "작년검색량",
    "작년최대검색월",
    "작년최대검색월검색량",
    "네이버해외상품수",
    "네이버경쟁강도",
    "네이버일반배송비율",
    "네이버묶음상품비율",
    "쿠팡평균가",
    "쿠팡총리뷰수",
    "쿠팡최대리뷰수",
    "쿠팡평균리뷰수",
    "쿠팡로켓배송비율",
    "쿠팡판매자로켓배송비율",
    "쿠팡일반배송비율",
    "쿠팡해외배송총리뷰수",
    "쿠팡해외배송최대리뷰수",
    "쿠팡해외배송평균리뷰수",
]

NAVER_PRODUCT_MAX = 30000
SHIPPING_MIN_RATE = 0.1      # 엑셀 내부값은 소수 (0.1 = 10%)
NAVER_PRICE_MAX = 5_000_000


def _num(v):
    """pd.to_numeric(errors='coerce') 대응. 숫자로 못 바꾸면 None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", ""))
    except (ValueError, TypeError):
        return None


def _ox(v):
    """'O'/'X' 판정용 정규화. 원본의 astype(str).strip().upper() 와 동일."""
    return str(v).strip().upper()


def process_one(file_path, out_path, ship_rate, max_price, max_products):
    """엑셀 1개를 스트리밍으로 필터링해 out_path 에 저장. (읽은행, 남은행) 반환."""
    fname = os.path.basename(file_path)
    try:
        wb_in = load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        print(f"  [읽기 실패] {fname} -> {e}")
        return None
    ws_in = wb_in.active

    rows = ws_in.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        print(f"  [건너뜀] {fname}: 빈 파일")
        wb_in.close()
        return None

    # 기준1: 남길 열 인덱스
    drop_norm = {norm(c) for c in COLS_TO_DELETE}
    keep_idx = [i for i, h in enumerate(header) if norm(h) not in drop_norm]

    # 필터에 쓸 원본 열 인덱스 (없으면 그 필터는 건너뜀)
    i_brand = find_col(header, "브랜드키워드")
    i_shop = find_col(header, "쇼핑성키워드")
    i_ncnt = find_col(header, "네이버상품수")
    i_nship = find_col(header, "네이버해외배송비율")
    i_cship = find_col(header, "쿠팡해외배송비율")
    i_price = find_col(header, "네이버평균가")

    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "Sheet1"  # detect.py 와의 계약
    ws_out.append([header[i] for i in keep_idx])

    before = after = 0
    for row in rows:
        if row is None:
            continue
        before += 1

        def cell(i):
            return row[i] if i is not None and i < len(row) else None

        # 기준2 / 기준3
        if i_brand is not None and _ox(cell(i_brand)) == "O":
            continue
        if i_shop is not None and _ox(cell(i_shop)) == "X":
            continue

        # 기준4: 숫자 아니면 제외 (원본 to_numeric coerce -> NaN <= max == False)
        if i_ncnt is not None:
            n = _num(cell(i_ncnt))
            if n is None or n > max_products:
                continue

        # 기준5 / 기준6: 숫자 아니면 제외
        if i_nship is not None:
            n = _num(cell(i_nship))
            if n is None or n < ship_rate:
                continue
        if i_cship is not None:
            n = _num(cell(i_cship))
            if n is None or n < ship_rate:
                continue

        # 기준7: 빈칸/0 은 가격정보 없음 -> 유지
        if i_price is not None:
            p = _num(cell(i_price))
            if p is not None and p != 0 and p > max_price:
                continue

        ws_out.append([cell(i) for i in keep_idx])
        after += 1

    wb_in.close()

    # 해외배송비율 열에 % 서식 복원
    out_header = [header[i] for i in keep_idx]
    for j, h in enumerate(out_header, start=1):
        if "해외배송비율" in norm(h):
            letter = get_column_letter(j)
            for r in range(2, after + 2):
                ws_out[f"{letter}{r}"].number_format = "0.00%"

    try:
        wb_out.save(out_path)
    except Exception as e:
        print(f"  [저장 실패] {os.path.basename(out_path)} -> {e}")
        return None

    dropped = len(header) - len(keep_idx)
    print(f"  [완료] {fname}: {before:,}행 -> {after:,}행 (열 {dropped}개 삭제, {len(keep_idx)}열 남음)")
    return before, after


def main():
    ap = argparse.ArgumentParser(description="셀러라이프 rawdata 필터")
    ap.add_argument("--in-dir", required=True, help="rawdata xlsx 폴더")
    ap.add_argument("--out-dir", required=True, help="filtered 저장 폴더")
    ap.add_argument("--ship-rate", type=float, default=SHIPPING_MIN_RATE,
                    help=f"해외배송비율 최소 (기본 {SHIPPING_MIN_RATE} = 10%%)")
    ap.add_argument("--max-price", type=float, default=NAVER_PRICE_MAX,
                    help=f"네이버 평균가 최대 (기본 {NAVER_PRICE_MAX:,})")
    ap.add_argument("--max-products", type=float, default=NAVER_PRODUCT_MAX,
                    help=f"네이버 상품수 최대 (기본 {NAVER_PRODUCT_MAX:,})")
    args = ap.parse_args()

    if not os.path.isdir(args.in_dir):
        print(f"[오류] 폴더 없음: {args.in_dir}")
        sys.exit(1)
    os.makedirs(args.out_dir, exist_ok=True)

    files = sorted(f for f in glob.glob(os.path.join(args.in_dir, "*.xlsx"))
                   if not os.path.basename(f).startswith("~$"))
    if not files:
        print(f"[알림] '{args.in_dir}' 안에 xlsx 가 없습니다.")
        sys.exit(1)

    print(f"총 {len(files)}개 파일 처리 (해외배송비율>={args.ship_rate}, "
          f"평균가<={args.max_price:,.0f}, 상품수<={args.max_products:,.0f})\n")

    ok, total_after = 0, 0
    for fp in files:
        name = os.path.basename(fp)
        out_path = os.path.join(args.out_dir, name if name.startswith("filtered_") else "filtered_" + name)
        try:
            res = process_one(fp, out_path, args.ship_rate, args.max_price, args.max_products)
        except Exception as e:
            print(f"  [처리 실패] {name} -> {e}")
            continue
        if res:
            ok += 1
            total_after += res[1]

    print(f"\n완료: {ok}/{len(files)}개 파일, 총 {total_after:,}행 -> {args.out_dir}")


if __name__ == "__main__":
    main()
