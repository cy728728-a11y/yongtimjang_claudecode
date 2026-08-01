#!/usr/bin/env python3
"""검토 엑셀에서 '승인' 표시된 키워드만 블랙리스트에 추가.

run_all.py 는 이 스크립트를 절대 호출하지 않는다.
이룸님이 검토_*.xlsx 의 '승인' 열에 O 를 찍고 저장한 뒤에만 실행한다.

  - 승인 값으로 인정: O o ㅇ 0 v V y Y ✓ (공백 무시)
  - 원본은 타임스탬프 백업(.bak_YYMMDD_HHMM.xlsx) 후 수정
  - '제외키워드' 시트에 (키워드, 카테고리, 제외사유) append
  - 이미 있는 키워드는 건너뜀

사용법:
  python blacklist_update.py --review-dir <run폴더>
  python blacklist_update.py --review-dir <run폴더> --dry-run   # 미리보기만
"""
import argparse
import glob
import os
import shutil
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook  # noqa: E402

from common import default_blacklist  # noqa: E402

SHEET_EXCLUDE_KW = "제외키워드"
# 'X' 는 일부러 뺀다 — 보통 '아님/반려' 의 뜻이라 승인으로 읽으면 위험하다.
APPROVE_MARKS = {"O", "0", "ㅇ", "V", "Y", "✓", "√"}


def _is_approved(v):
    if v is None:
        return False
    s = str(v).strip().upper()
    return bool(s) and s in APPROVE_MARKS


def collect_approved(review_dir):
    """검토_*.xlsx 에서 승인 표시된 (키워드, 카테고리, 사유) 수집."""
    files = sorted(glob.glob(os.path.join(review_dir, "검토_*.xlsx")))
    if not files:
        print(f"[오류] '{review_dir}' 안에 검토_*.xlsx 가 없습니다.")
        return None

    items, total_rows = [], 0
    for fp in files:
        wb = load_workbook(fp, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [str(h or "").strip() for h in next(it)]
        try:
            i_kw = header.index("키워드")
            i_cat = header.index("카테고리")
            i_reason = header.index("AI사유")
            i_ok = header.index("승인")
        except ValueError:
            print(f"  [건너뜀] {os.path.basename(fp)}: 헤더가 예상과 다릅니다 ({header})")
            wb.close()
            continue

        n = 0
        for row in it:
            if row is None or row[i_kw] is None:
                continue
            total_rows += 1
            if _is_approved(row[i_ok]):
                items.append((str(row[i_kw]).strip(),
                              str(row[i_cat] or "").strip(),
                              str(row[i_reason] or "").strip()))
                n += 1
        wb.close()
        print(f"  {os.path.basename(fp)}: 승인 {n}건")

    print(f"\n검토 대상 {total_rows:,}건 중 승인 {len(items)}건")
    return items


def append_keywords(blacklist_path, items, dry_run=False):
    if not os.path.exists(blacklist_path):
        print(f"[오류] 블랙리스트 없음: {blacklist_path}")
        return 0

    wb = load_workbook(blacklist_path)
    if SHEET_EXCLUDE_KW not in wb.sheetnames:
        print(f"[오류] '{SHEET_EXCLUDE_KW}' 시트 없음")
        return 0
    ws = wb[SHEET_EXCLUDE_KW]

    existing = set()
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if row[0] is not None and str(row[0]).strip():
            existing.add(str(row[0]).strip())

    new_items = [it for it in items if it[0] not in existing]
    dup = len(items) - len(new_items)
    if dup:
        print(f"이미 블랙리스트에 있음 (건너뜀): {dup}건")

    if not new_items:
        print("추가할 키워드가 없습니다.")
        return 0

    print(f"\n추가될 {len(new_items)}건:")
    for kw, cat, reason in new_items[:20]:
        print(f"  + {kw}  ({reason})")
    if len(new_items) > 20:
        print(f"  ... 외 {len(new_items) - 20}건")

    if dry_run:
        print("\n[dry-run] 실제로 쓰지 않았습니다.")
        return 0

    stamp = datetime.now().strftime("%y%m%d_%H%M")
    backup = f"{os.path.splitext(blacklist_path)[0]}.bak_{stamp}.xlsx"
    shutil.copy2(blacklist_path, backup)
    print(f"\n백업: {backup}")

    for kw, cat, reason in new_items:
        ws.append([kw, cat, reason])
    wb.save(blacklist_path)
    print(f"블랙리스트에 {len(new_items)}건 추가 -> {blacklist_path}")
    return len(new_items)


def main():
    ap = argparse.ArgumentParser(description="승인된 AI위험 키워드를 블랙리스트에 추가")
    ap.add_argument("--review-dir", required=True, help="검토_*.xlsx 가 있는 폴더")
    ap.add_argument("--blacklist", help="기본: <DATA_ROOT>/keyword_blacklist/keyword_blacklist.xlsx")
    ap.add_argument("--dry-run", action="store_true", help="미리보기만, 쓰지 않음")
    args = ap.parse_args()

    items = collect_approved(args.review_dir)
    if items is None:
        sys.exit(1)
    if not items:
        print("승인 표시된 키워드가 없습니다. 검토_*.xlsx 의 '승인' 열에 O 를 찍고 저장하세요.")
        return

    append_keywords(args.blacklist or default_blacklist(), items, args.dry_run)


if __name__ == "__main__":
    main()
