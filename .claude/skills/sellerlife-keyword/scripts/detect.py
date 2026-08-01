#!/usr/bin/env python3
"""브랜드/상표/캐릭터 키워드 판정 (블랙리스트 + Gemini).

gemini_detector.py 포팅:
  - API 키/경로 하드코딩 제거 -> .env + CLI
  - 키워드/카테고리 열을 고정 인덱스(B=2,C=3) 대신 헤더명으로 해석
  - google-genai SDK 대신 requests REST (의존성 0, AQ./AIza 키 모두 동작 확인)
  - risk 를 스키마 enum('Y','N')으로 고정 (모델이 'low' 같은 값을 뱉지 못하게)

처리 순서 (행마다):
  1단계 카테고리 제외 : 카테고리에 '제외카테고리' 조각이 포함되면 제외
  2단계 키워드 정확일치 제외 : 키워드가 '제외키워드'와 정확히 같으면 제외
  2.5단계 브랜드 부분일치 제외 : '제외브랜드' 어근이 키워드에 '포함'되면 제외 (예: '하리오'→'하리오알파02')
  3단계 AI 판정 : 남은 키워드만 Gemini 배치 전송

이어하기(RESUME): 결과파일(_AI검출.xlsx)이 이미 있으면 그것을 열어
'판정'이 비었거나 'AI오류-재실행필요'인 행만 다시 처리한다.

사용법:
  python detect.py --in-dir <filtered폴더> [--blacklist <경로>] [--batch-size 60]
"""
import argparse
import glob
import json
import os
import re
import sys
import time

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import requests  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from common import default_blacklist, load_env, norm  # noqa: E402

INPUT_SHEET = "Sheet1"
SHEET_EXCLUDE_KW = "제외키워드"
SHEET_EXCLUDE_CAT = "제외카테고리"
SHEET_EXCLUDE_BRAND = "제외브랜드"  # 부분일치: 어근이 키워드에 '포함'되면 제외 (브랜드/모델명)

MODEL = "gemini-2.5-flash-lite"
BATCH_SIZE = 60
SLEEP_SEC = 1
MAX_RETRY = 4
RETRY_BASE_WAIT = 5
SAVE_EVERY = 20
MAX_EXAMPLE = 80

AI_COLS = ["상태", "AI위험판정", "AI분류", "AI사유", "판정"]
DONE_JUDGES = ("제외", "AI위험", "안전")

API_URL = "https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"

RESPONSE_SCHEMA = {
    "type": "ARRAY",
    "items": {
        "type": "OBJECT",
        "properties": {
            "index": {"type": "INTEGER"},
            "risk": {"type": "STRING", "enum": ["Y", "N"]},
            "type": {"type": "STRING"},
            "reason": {"type": "STRING"},
        },
        "required": ["index", "risk", "type", "reason"],
    },
}


# ---------------------------------------------------------------- 블랙리스트
def load_blacklist(path):
    """(제외키워드 set, 제외카테고리 list, [(키워드,사유)] list, 제외브랜드 list) 반환.

    제외브랜드는 부분일치용 어근 목록(소문자화). 키워드에 어근이 '포함'되면 제외.
    """
    if not os.path.exists(path):
        print(f"[오류] 기준파일 없음: {path}")
        return None, None, None, None

    wb = load_workbook(path, read_only=True)
    kw_set, reasoned, cat_list, brand_list = set(), [], [], []

    if SHEET_EXCLUDE_KW in wb.sheetnames:
        # A=키워드, C=제외사유. 1행은 헤더.
        for row in wb[SHEET_EXCLUDE_KW].iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
            kw = row[0]
            if kw is None or not str(kw).strip():
                continue
            k = str(kw).strip()
            kw_set.add(k)
            reason = row[2] if len(row) >= 3 else None
            if reason is not None and str(reason).strip():
                reasoned.append((k, str(reason).strip()))
    else:
        print(f"[경고] '{SHEET_EXCLUDE_KW}' 시트 없음 - 키워드 제외 건너뜀")

    if SHEET_EXCLUDE_CAT in wb.sheetnames:
        # 헤더 없음 -> 1행부터
        for row in wb[SHEET_EXCLUDE_CAT].iter_rows(min_col=1, max_col=1, values_only=True):
            v = row[0]
            if v is not None and str(v).strip():
                cat_list.append(str(v).strip())
    else:
        print(f"[경고] '{SHEET_EXCLUDE_CAT}' 시트 없음 - 카테고리 제외 건너뜀")

    if SHEET_EXCLUDE_BRAND in wb.sheetnames:
        # A=브랜드어근(부분일치). 1행은 헤더. 소문자화해 대소문자 무시.
        for row in wb[SHEET_EXCLUDE_BRAND].iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
            v = row[0]
            if v is not None and str(v).strip():
                brand_list.append(str(v).strip().lower())
    # 시트가 없어도 조용히 넘어감(선택 기능)

    wb.close()
    return kw_set, cat_list, reasoned, brand_list


# ---------------------------------------------------------------- Gemini
def build_prompt(batch, examples, reasoned_examples):
    ex_text = ", ".join(examples[:MAX_EXAMPLE])
    items = "\n".join(f"{i}: {kw}" for i, kw in enumerate(batch))

    reason_block = ""
    if reasoned_examples:
        lines = "\n".join(f"- {kw} → {reason}" for kw, reason in reasoned_examples[:40])
        reason_block = f"""
[셀러가 '이런 이유로 제외한다'고 밝힌 사례 - 같은 맥락의 키워드도 위험(Y)으로 판정하라]
{lines}
"""
    return f"""너는 네이버 스마트스토어 해외구매대행 셀러를 돕는 키워드 분류기다.
아래 검색 키워드 각각을, 구매대행으로 팔기 부적절하여 '걸러내야' 하는지 판정하라.
걸러내야 하는 경우는 다음과 같다:
(1) 브랜드명, 상표권 있는 제품명, 캐릭터/애니/IP 명칭, 유명 디자이너/제품 라인을 포함 → 상표권 침해 위험
(2) 아래 '셀러가 밝힌 제외 사유'와 같은 맥락에 해당하는 경우 (예: 어린이용 제품 등)
'2인용소파', '원목책상', '접이식테이블'처럼 일반 상품명이고 위 사유에도 해당 안 되면 안전(N)이다.
글자가 한글이라도 외국 브랜드/캐릭터의 음차 표기일 수 있으니 의미로 판단하라.

[셀러가 직접 '위험'으로 분류해 둔 예시]
{ex_text}
{reason_block}
[판정 대상] (번호: 키워드)
{items}

각 키워드에 대해 아래 형식의 JSON 배열만 출력하라. 설명·마크다운 금지.
index=위 번호, risk="Y"(위험) 또는 "N"(안전), type=브랜드/캐릭터/상표/디자이너/어린이용/기타/일반 중 하나, reason=10자 내외 짧은 근거."""


def classify_batch(api_key, model, batch, examples, reasoned_examples):
    """Gemini REST 호출. {local_index: {risk,type,reason}} 반환. 실패 시 예외."""
    body = {
        "contents": [{"parts": [{"text": build_prompt(batch, examples, reasoned_examples)}]}],
        "generationConfig": {
            "responseMimeType": "application/json",
            "responseSchema": RESPONSE_SCHEMA,
            "temperature": 0,
        },
    }
    r = requests.post(API_URL.format(model=model),
                      headers={"x-goog-api-key": api_key}, json=body, timeout=120)
    if r.status_code != 200:
        raise RuntimeError(f"HTTP {r.status_code}: {r.text[:200]}")

    text = r.json()["candidates"][0]["content"]["parts"][0]["text"].strip()
    text = re.sub(r"^```(json)?|```$", "", text).strip()
    data = json.loads(text)

    out = {}
    for item in data:
        out[int(item["index"])] = {
            "risk": str(item.get("risk", "")).upper(),
            "type": str(item.get("type", "")),
            "reason": str(item.get("reason", "")),
        }
    return out


def probe_key(api_key, model):
    """배치 루프 전에 키/모델이 실제로 동작하는지 1건으로 확인."""
    try:
        classify_batch(api_key, model, ["2인용소파"], [], [])
        return True
    except Exception as e:
        print(f"[오류] Gemini 호출 실패: {e}")
        print("  → .env 의 GEMINI_API_KEY 를 확인하세요 (https://aistudio.google.com/apikey).")
        return False


# ---------------------------------------------------------------- 엑셀
def resolve_cols(ws):
    """헤더에서 키워드/카테고리 열 번호(1-based)를 찾는다. 없으면 기존 고정값(2,3) 폴백."""
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    kw = cat = None
    for i, h in enumerate(header, start=1):
        n = norm(h)
        if n == "키워드" and kw is None:
            kw = i
        elif n == "카테고리" and cat is None:
            cat = i
    if kw is None or cat is None:
        print(f"[경고] 헤더에서 키워드/카테고리 열을 못 찾아 기본값(2,3) 사용")
        return kw or 2, cat or 3
    return kw, cat


def find_or_make_columns(ws):
    """AI 결과 열 5개의 위치를 찾거나 오른쪽에 새로 만든다."""
    existing = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v in AI_COLS:
            existing[v] = c
    if len(existing) == len(AI_COLS):
        return [existing[n] for n in AI_COLS]

    base = ws.max_column
    cols = list(range(base + 1, base + 1 + len(AI_COLS)))
    for col, name in zip(cols, AI_COLS):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, w in zip(cols, [16, 11, 12, 30, 14]):
        ws.column_dimensions[get_column_letter(col)].width = w
    return cols


def process_file(in_path, api_key, model, kw_set, cat_list, reasoned, examples,
                 batch_size, sleep_sec, brand_list=()):
    print(f"\n{'='*60}\n■ {os.path.basename(in_path)}\n{'='*60}")

    out_path = os.path.splitext(in_path)[0] + "_AI검출.xlsx"
    resume = os.path.exists(out_path)
    src = out_path if resume else in_path
    print(("이어하기: 오류/미처리 행만 재실행" if resume else "새로 시작") + f"  <- {os.path.basename(src)}")

    wb = load_workbook(src)
    if INPUT_SHEET not in wb.sheetnames:
        print(f"[오류] '{INPUT_SHEET}' 시트 없음. 시트: {wb.sheetnames}")
        return 0
    ws = wb[INPUT_SHEET]

    kw_col, cat_col = resolve_cols(ws)
    c_state, c_risk, c_type, c_reason, c_judge = find_or_make_columns(ws)

    fill_cat = PatternFill("solid", fgColor="D9D9D9")
    fill_kw = PatternFill("solid", fgColor="FFD966")
    fill_brand = PatternFill("solid", fgColor="F4B183")
    fill_risk = PatternFill("solid", fgColor="FFC000")
    fill_none = PatternFill(fill_type=None)

    to_ai = []
    n_cat = n_kw = n_brand = n_done = 0
    for r in range(2, ws.max_row + 1):
        if resume and ws.cell(row=r, column=c_judge).value in DONE_JUDGES:
            n_done += 1
            continue

        kw = ws.cell(row=r, column=kw_col).value
        cat = ws.cell(row=r, column=cat_col).value
        kw_s = "" if kw is None else str(kw).strip()
        cat_s = "" if cat is None else str(cat)

        # 1단계) 카테고리 제외
        if cat_list and any(frag in cat_s for frag in cat_list):
            ws.cell(row=r, column=c_state, value="카테고리제외").fill = fill_cat
            ws.cell(row=r, column=c_judge, value="제외")
            n_cat += 1
            continue

        # 2단계) 키워드 정확일치 제외
        if kw_s in kw_set:
            ws.cell(row=r, column=c_state, value="키워드일치제외").fill = fill_kw
            ws.cell(row=r, column=c_judge, value="제외")
            n_kw += 1
            continue

        # 2.5단계) 브랜드 어근 부분일치 제외 (제외브랜드 시트)
        if brand_list:
            kw_low = kw_s.lower()
            hit = next((b for b in brand_list if b in kw_low), None)
            if hit:
                ws.cell(row=r, column=c_state, value=f"브랜드제외({hit})").fill = fill_brand
                ws.cell(row=r, column=c_judge, value="제외")
                n_brand += 1
                continue

        # 3단계) AI 판정 대상 (오류 행은 표시 초기화)
        ws.cell(row=r, column=c_state, value="AI판정대상")
        jc = ws.cell(row=r, column=c_judge, value=None)
        jc.fill = fill_none
        to_ai.append((r, kw_s))

    if resume:
        print(f"이미 완료 {n_done}개 / 이번 AI대상 {len(to_ai)}개")
    print(f"제외: 카테고리 {n_cat} + 키워드일치 {n_kw} + 브랜드부분일치 {n_brand} / AI판정 대상 {len(to_ai)}\n")

    if not to_ai:
        wb.save(out_path)
        print(f"AI 대상 없음. 저장: {out_path}")
        return 0

    batches = [to_ai[i:i + batch_size] for i in range(0, len(to_ai), batch_size)]
    n_risk = n_safe = n_err = 0

    for bi, batch in enumerate(batches, start=1):
        kw_list = [kw for _r, kw in batch]
        result = None
        for attempt in range(1, MAX_RETRY + 1):
            try:
                result = classify_batch(api_key, model, kw_list, examples, reasoned)
                break
            except Exception as e:
                msg = str(e)
                short = "한도초과(429)" if "429" in msg or "RESOURCE_EXHAUSTED" in msg else msg[:80]
                print(f"  [배치 {bi}] 시도 {attempt} 실패: {short}")
                if attempt < MAX_RETRY:
                    time.sleep(RETRY_BASE_WAIT * attempt)

        for local_idx, (row_no, _kw) in enumerate(batch):
            jcell = ws.cell(row=row_no, column=c_judge)
            if result and local_idx in result:
                v = result[local_idx]
                ws.cell(row=row_no, column=c_risk, value=v["risk"])
                ws.cell(row=row_no, column=c_type, value=v["type"])
                ws.cell(row=row_no, column=c_reason, value=v["reason"])
                if v["risk"] == "Y":
                    jcell.value = "AI위험"
                    jcell.fill = fill_risk
                    n_risk += 1
                else:
                    jcell.value = "안전"
                    n_safe += 1
            else:
                jcell.value = "AI오류-재실행필요"
                n_err += 1

        print(f"  배치 {bi}/{len(batches)} (위험 {n_risk} / 안전 {n_safe} / 오류 {n_err})")
        if bi % SAVE_EVERY == 0:
            wb.save(out_path)
            print(f"    ...중간 저장")
        time.sleep(sleep_sec)

    wb.save(out_path)
    print(f"=== 완료: AI위험 {n_risk} / 안전 {n_safe} / 오류 {n_err}")
    print(f"    저장: {out_path}")
    if n_err:
        print("    ※ 같은 명령을 한 번 더 실행하면 오류 행만 재처리됩니다.")
    return n_err


def main():
    ap = argparse.ArgumentParser(description="블랙리스트 + Gemini 키워드 판정")
    ap.add_argument("--in-dir", required=True, help="filtered xlsx 폴더")
    ap.add_argument("--out-dir", help="결과 저장 폴더 (기본: --in-dir 과 동일)")
    ap.add_argument("--blacklist", help="기본: <DATA_ROOT>/keyword_blacklist/keyword_blacklist.xlsx")
    ap.add_argument("--model", default=MODEL)
    ap.add_argument("--batch-size", type=int, default=BATCH_SIZE)
    ap.add_argument("--sleep", type=float, default=SLEEP_SEC)
    args = ap.parse_args()

    cfg = load_env()
    api_key = cfg.get("GEMINI_API_KEY")
    if not api_key:
        print("[오류] GEMINI_API_KEY 없음. .env 를 확인하세요.")
        sys.exit(1)

    if not os.path.isdir(args.in_dir):
        print(f"[오류] 폴더 없음: {args.in_dir}")
        sys.exit(1)

    files = sorted(f for f in glob.glob(os.path.join(args.in_dir, "*.xlsx"))
                   if not os.path.basename(f).startswith("~$")
                   and not os.path.basename(f).endswith("_AI검출.xlsx"))
    if not files:
        print(f"[알림] '{args.in_dir}' 안에 처리할 엑셀이 없습니다.")
        sys.exit(1)

    kw_set, cat_list, reasoned, brand_list = load_blacklist(
        args.blacklist or default_blacklist())
    if kw_set is None:
        sys.exit(1)
    print(f"기준파일: 제외키워드 {len(kw_set)}개 / 제외카테고리 {len(cat_list)}개 / "
          f"제외브랜드(부분일치) {len(brand_list)}개 / 사유입력 {len(reasoned)}개")
    examples = sorted(kw_set)

    if not probe_key(api_key, args.model):
        sys.exit(1)
    print(f"Gemini 연결 확인 ({args.model})")

    print(f"\n총 {len(files)}개 파일 처리 시작")
    total_err = 0
    for fp in files:
        try:
            total_err += process_file(fp, api_key, args.model, kw_set, cat_list, reasoned,
                                      examples, args.batch_size, args.sleep,
                                      brand_list=brand_list) or 0
        except Exception as e:
            print(f"[파일 처리 실패] {os.path.basename(fp)} -> {e}")

    print(f"\n{'#'*60}\n전체 완료: {len(files)}개 파일 / 남은 오류 {total_err}개")
    if total_err:
        print("오류가 남았으면 같은 명령으로 다시 실행하세요 (오류 행만 재처리).")


if __name__ == "__main__":
    main()
