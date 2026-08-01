#!/usr/bin/env python3
"""다운로드 산출물을 카테고리별 rawdata 엑셀로 정규화.

셀러라이프 '전체 카테고리 엑셀 다운로드' 는 ZIP(카테고리별 엑셀) 이지만,
단일 엑셀(카테고리 열 포함) 로 바뀌어도 동작하도록 양쪽 다 처리한다.

  zip    -> 멤버 중 대상 카테고리(가구/인테리어·디지털/가전·생활/건강)만 추출
  single -> '카테고리' 열의 1차 분류로 행을 갈라 카테고리별 파일 생성

'생활&건강' 과 '생활&건강2' 처럼 20만행 export 상한 때문에 쪼개진 파일은
같은 카테고리로 인식되어 뒤 단계(finalize)에서 하나로 병합된다.
여기서는 파일을 그대로 두되 이름만 정규화한다.

사용법:
  python ingest.py --artifact <받은파일> --out-dir <raw폴더>
"""
import argparse
import json
import os
import sys
import zipfile

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import Workbook, load_workbook  # noqa: E402

from common import KEEP_CATEGORIES, KEEP_KEYS, cat_key, find_col  # noqa: E402


def _decode_member(name):
    """ZIP 멤버명 한글 깨짐 복구 (cp437 로 잘못 디코드된 경우)."""
    try:
        return name.encode("cp437").decode("cp949")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name


def from_zip(artifact, out_dir):
    kept = []
    with zipfile.ZipFile(artifact) as z:
        for info in z.infolist():
            if info.is_dir():
                continue
            raw_name = info.filename
            name = _decode_member(raw_name) if not (info.flag_bits & 0x800) else raw_name
            base = os.path.basename(name)
            if not base.lower().endswith((".xlsx", ".xls")):
                continue

            stem = os.path.splitext(base)[0]
            # 'SellarSourcing_생활&건강2' -> '생활건강'
            label = stem.split("_", 1)[1] if "_" in stem else stem
            ck = cat_key(label)

            if ck not in KEEP_KEYS:
                print(f"  [제외] {base}  (대상 카테고리 아님: {ck})")
                continue

            dest = os.path.join(out_dir, base)
            with z.open(info) as src, open(dest, "wb") as out:
                out.write(src.read())
            size_mb = os.path.getsize(dest) / 1048576
            print(f"  [추출] {base}  ({ck}, {size_mb:.1f} MB)")
            kept.append(dest)
    return kept


def from_single(artifact, out_dir):
    """카테고리 열 기준으로 행을 갈라 카테고리별 파일 생성 (스트리밍)."""
    wb = load_workbook(artifact, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = list(next(it))

    i_cat = find_col(header, "카테고리")
    if i_cat is None:
        wb.close()
        raise RuntimeError("단일 엑셀인데 '카테고리' 열이 없습니다. --probe 결과를 확인하세요.")

    writers = {}
    counts = {}
    for row in it:
        if row is None or i_cat >= len(row):
            continue
        ck = cat_key(row[i_cat])
        if ck not in KEEP_KEYS:
            continue
        if ck not in writers:
            w = Workbook()
            s = w.active
            s.title = "all"
            s.append(header)
            writers[ck] = (w, s)
            counts[ck] = 0
        writers[ck][1].append(list(row))
        counts[ck] += 1
    wb.close()

    kept = []
    for ck, (w, _s) in writers.items():
        dest = os.path.join(out_dir, f"SellarSourcing_{ck}.xlsx")
        w.save(dest)
        print(f"  [분리] {os.path.basename(dest)}  ({counts[ck]:,}행)")
        kept.append(dest)
    return kept


def normalize(artifact, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    print(f"대상 카테고리: {', '.join(KEEP_CATEGORIES)}\n")

    if zipfile.is_zipfile(artifact):
        print("산출물 형태: ZIP")
        kept = from_zip(artifact, out_dir)
    else:
        print("산출물 형태: 단일 엑셀")
        kept = from_single(artifact, out_dir)

    if not kept:
        raise RuntimeError("대상 카테고리 파일을 하나도 찾지 못했습니다. "
                           "카테고리 표기가 바뀌었는지 확인하세요.")
    return kept


def main():
    ap = argparse.ArgumentParser(description="다운로드 산출물 -> 카테고리별 rawdata")
    ap.add_argument("--artifact", required=True, help="download.py 가 받은 zip 또는 xlsx")
    ap.add_argument("--out-dir", required=True, help="raw 저장 폴더")
    args = ap.parse_args()

    if not os.path.exists(args.artifact):
        print(f"[오류] 파일 없음: {args.artifact}")
        sys.exit(1)

    kept = normalize(args.artifact, args.out_dir)
    print(f"\n총 {len(kept)}개 카테고리 파일 준비됨")
    print("###RAWFILES###")
    print(json.dumps(kept, ensure_ascii=False))


if __name__ == "__main__":
    main()
