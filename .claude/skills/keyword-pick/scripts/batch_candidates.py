#!/usr/bin/env python3
"""프리필터 결과(카테고리별 xlsx) → 상품별 계단 태깅 후보 JSON 청크 생성.

키워드→상품명 배치 파이프라인 3단계 스크립트.
kwlib.build_candidates()로 후보를 뽑고, 상품수 구간(1만/2만/3만)으로 계단 태깅한 뒤
캡을 적용해 상품당 최대 35개(15+10+10)로 줄인다. 여기서 브랜드/직결여부 판정은 하지 않는다
(그건 Claude가 chunk를 읽고 판단). 이 스크립트는 기계적 준비만 한다.

입력 (run_dir 안에):
  - manifest.json : {"categories": {"<카테고리>": {"file": "prefiltered/xxx.xlsx"}}}
    file 경로는 run_dir 기준 상대경로 또는 절대경로 모두 허용.
  - targets.json  : [{"productId":"...", "상품명":"...", "대표키워드":"...", "카테고리":"..."}, ...]

출력:
  - <run_dir>/candidates/chunk_001.json, chunk_002.json, ... (상품 10개/청크, 경계로만 자름)

Usage:
  python batch_candidates.py <run_dir> [--chunk-size 10] [--tiers SPEC] [--blacklist PATH]

--tiers SPEC: 계단 구간을 "상한:캡" 쌍의 쉼표 목록으로 지정. 상한 0 = 무제한(마지막에만).
  기본값 "10000:15,20000:10,30000:10" = 기존 동작(④ keyword-pick 6개 기준).
  ⑤ product-name(키워드 2개 확보까지 무한 확장)은 더 긴 스펙을 넘긴다.
"""

import argparse
import json
import os
import sys
from pathlib import Path

# kwlib.py가 같은 폴더에 있으므로 어느 경로에서 실행해도 import 가능하도록 경로 추가
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from kwlib import build_candidates, extract_roots, load_rows  # noqa: F401 (extract_roots 재export)

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl>=3.1.0", file=sys.stderr)
    sys.exit(1)

DEFAULT_BLACKLIST = r"D:\python_work\data\sellerlife\keyword_blacklist\keyword_blacklist.xlsx"

# 계단 구간 상한(상품수) 및 구간별 캡 — 기본값 = ④ keyword-pick 기존 동작
DEFAULT_TIERS = "10000:15,20000:10,30000:10"
# 무제한 구간(상한 0)을 다룰 때 build_candidates에 넘길 실질 무한대
UNLIMITED = 10 ** 12


def parse_tiers(spec):
    """'10000:15,20000:10,0:8' -> [(10000,15),(20000,10),(0,8)]. 상한 0 = 무제한(마지막만 허용)."""
    tiers = []
    for part in str(spec).split(","):
        part = part.strip()
        if not part:
            continue
        if ":" not in part:
            raise ValueError(f"--tiers 형식 오류('상한:캡' 이어야 함): {part}")
        hi_s, cap_s = part.split(":", 1)
        try:
            hi, cap = int(hi_s), int(cap_s)
        except ValueError:
            raise ValueError(f"--tiers 숫자 파싱 실패: {part}")
        if cap <= 0:
            raise ValueError(f"--tiers 캡은 1 이상이어야 합니다: {part}")
        tiers.append((hi, cap))
    if not tiers:
        raise ValueError("--tiers 가 비어 있습니다.")
    for hi, _ in tiers[:-1]:
        if hi <= 0:
            raise ValueError("무제한 구간(상한 0)은 마지막에만 올 수 있습니다.")
    return tiers


def tier_label(hi, is_first):
    """구간 상한 -> 시트 '기준' 열에 쓸 라벨.
    첫 구간은 기본 컷이라 '확장' 표기 없음, 이후 구간은 '확장'을 붙인다(기존 시트 표기 유지)."""
    if hi <= 0:
        return "상품수 무제한 확장"
    base = f"상품수 {hi // 10000}만 이하" if hi % 10000 == 0 else f"상품수 {hi:,} 이하"
    return base if is_first else f"{base} 확장"

# 최종 출력에 남길 후보 필드(스키마 고정)
CANDIDATE_FIELDS = [
    "키워드", "총검색수", "상품수", "기준",
    "네이버해외배송비율", "쿠팡해외배송비율", "신규진입키워드", "브랜드의심",
]


def load_blacklist(path):
    """sellerlife 키워드 블랙리스트에서 제외키워드(정확일치)·제외브랜드(부분일치)를 읽는다.
    파일/시트가 없으면 조용히 빈 값 반환(설계문서 규칙)."""
    excluded_keywords = set()
    brand_roots = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return excluded_keywords, brand_roots

    try:
        if "제외키워드" in wb.sheetnames:
            ws = wb["제외키워드"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(h).strip() if h is not None else "" for h in rows[0]]
                kw_idx = header.index("키워드") if "키워드" in header else 0
                for r in rows[1:]:
                    if kw_idx < len(r) and r[kw_idx] is not None and str(r[kw_idx]).strip():
                        excluded_keywords.add(str(r[kw_idx]).strip())

        if "제외브랜드" in wb.sheetnames:
            ws = wb["제외브랜드"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(h).strip() if h is not None else "" for h in rows[0]]
                br_idx = header.index("브랜드어근") if "브랜드어근" in header else 0
                for r in rows[1:]:
                    if br_idx < len(r) and r[br_idx] is not None and str(r[br_idx]).strip():
                        brand_roots.append(str(r[br_idx]).strip())
    except Exception:
        # 시트 구조가 예상과 달라도 배치 전체를 막지 않는다 — 플래그 없이 진행
        pass
    finally:
        wb.close()

    return excluded_keywords, brand_roots


def is_brand_suspect(keyword, excluded_keywords, brand_roots):
    """블랙리스트 힌트 판정. 정확일치(제외키워드) 또는 부분일치(브랜드어근)면 True."""
    if keyword in excluded_keywords:
        return True
    for root in brand_roots:
        if root and root in keyword:
            return True
    return False


def pick_relevant(all_candidates, roots, cap, noise_ratio=0.6, min_roots=1):
    """상품 어근을 포함하는 후보를 최대 cap개 고른다.

    min_roots: 후보가 포함해야 할 최소 어근 수. leaf 풀은 1로 충분하지만
    상위 카테고리 풀은 2를 요구한다 — 범용 어근 1개('자동'·'온실'·'체어')만 걸리면
    전혀 다른 상품이 통째로 딸려온다(실측: '자동 접지선' → 자동바·자동톱·자동면).

    어근을 **구체적인 것 우선**(매칭 수가 적은 순)으로 훑는다. '자동' 같은 범용어보다
    '디스펜서' 같은 특정어가 먼저 자리를 차지하게 하려는 것. 어근 내부는 상품수 오름차순.
    전체의 noise_ratio를 넘게 걸리는 어근만 제외하되, 그 결과 한 건도 못 건지면
    제외했던 어근을 다시 쓴다(카테고리 전용어가 통째로 날아가는 것 방지).

    캡 구간에서 잘려나가는 직결 키워드를 살리는 게 목적이라 상품수 상한은 두지 않는다.
    """
    if cap <= 0 or not roots or not all_candidates:
        return []
    total = len(all_candidates)
    limit = max(1, int(total * noise_ratio))

    hit_counts = {}
    for r in roots:
        n = sum(1 for c in all_candidates if r in c["키워드"])
        if n:
            hit_counts[r] = n
    if not hit_counts:
        return []
    usable = {r for r, n in hit_counts.items() if n <= limit} or set(hit_counts)

    scored = []
    for c in all_candidates:
        kw = c["키워드"]
        nroots = sum(1 for r in usable if r in kw)
        if nroots < min_roots:
            continue
        # 정렬 우선순위:
        #  ① 상품 어근을 여러 개 포함할수록 직결일 가능성이 높다
        #  ② 짧을수록 일반명사(브랜드 키워드는 '브랜드+명사'라 길어진다)
        #  ③ 브랜드의심 플래그는 뒤로
        #  ④ 같은 조건이면 상품수 낮은 것
        scored.append(((-nroots, 1 if c.get("브랜드의심") else 0, len(kw), c["상품수"]), c))

    out, seen = [], set()
    for _, c in sorted(scored, key=lambda x: x[0]):
        if c["키워드"] in seen:
            continue
        seen.add(c["키워드"])
        out.append(c)
        if len(out) >= cap:
            break
    return out


def _prefix(cat, n):
    """카테고리 경로의 앞 n개 세그먼트를 공백 제거해 반환. 접두 일치 비교용."""
    parts = [p.strip() for p in str(cat or "").split(">")]
    return ">".join(parts[:n]).replace(" ", "")


def pick_parent_relevant(parent_pool, roots, cap, product_category, exclude=(), min_roots=2):
    """상위 카테고리 풀에서 상품 어근에 걸리는 후보를 최대 cap개 고른다.

    - leaf 후보에 이미 있는 키워드는 **선별 전에** 뺀다. 안 그러면 중복이 cap을 먹고
      병합 단계에서 버려져 자리만 낭비된다(실측: 자키 10칸 중 4칸이 그렇게 증발).
    - 3차 근접(같은 3차 접두)은 **선별 순서에 쓰지 않고 `기준` 표기로만 남긴다.**
      실측에서 자키의 진짜 직결어(전동유압자키·유압자키)는 3차 형제인 '운반용품'이 아니라
      2차 건너편 '에어공구>유압공구'에 있었다. 근접 차수는 적합성의 대리지표가 못 된다.
    """
    if cap <= 0 or not parent_pool:
        return []
    ex = set(exclude)
    pool = [c for c in parent_pool if c["키워드"] not in ex]
    p3 = _prefix(product_category, 3)

    out = []
    for c in pick_relevant(pool, roots, cap, min_roots=min_roots):
        picked = dict(c)
        picked["기준"] = ("상위카테고리(3차)"
                          if _prefix(c.get("대표카테고리"), 3) == p3
                          else "상위카테고리(2차)")
        out.append(picked)
    return out


def build_capped_candidates(xlsx_path, excluded_keywords, brand_roots, tiers):
    """카테고리 xlsx 하나에서 계단 태깅 + 캡 적용된 후보 목록을 만든다.

    tiers = [(상한, 캡), ...] 오름차순. 상한 0 = 무제한(마지막 구간).
    구간별로 상품수 오름차순 상위 <캡>개만 남긴다 — 낮은 구간부터 채택하는 게 원칙이므로
    Claude는 앞쪽(낮은 상품수) 후보부터 보게 된다.
    """
    idx, data = load_rows(xlsx_path)
    ceiling = tiers[-1][0]
    all_candidates = build_candidates(
        idx, data, UNLIMITED if ceiling <= 0 else ceiling, exclude_info=True
    )

    for c in all_candidates:
        c["브랜드의심"] = is_brand_suspect(c["키워드"], excluded_keywords, brand_roots)

    all_candidates.sort(key=lambda c: c["상품수"])

    capped = []
    low = -1  # 상품수 0인 키워드도 첫 구간에 포함되도록
    for i, (hi, cap) in enumerate(tiers):
        top = UNLIMITED if hi <= 0 else hi
        bucket = [c for c in all_candidates if low < c["상품수"] <= top]
        label = tier_label(hi, is_first=(i == 0))
        for c in bucket:      # 캡 밖 후보도 라벨을 달아둔다(관련어 선별에서 쓰임)
            c["기준"] = label
        capped.extend(bucket[:cap])
        low = top

    # 스키마 고정 필드만 남긴다 (청크 용량 절약)
    slim = [{k: c.get(k) for k in CANDIDATE_FIELDS} for c in capped]
    return slim, all_candidates  # all_candidates = 계단 태깅된 전체 풀(상품별 관련어 선별용)


def main():
    parser = argparse.ArgumentParser(description="프리필터 결과 → 상품별 계단태깅 후보 JSON 청크 생성")
    parser.add_argument("run_dir", help="manifest.json/targets.json이 있는 실행 디렉토리. candidates/ 하위에 출력")
    parser.add_argument("--chunk-size", type=int, default=10, help="청크당 상품 수 (default: 10)")
    parser.add_argument("--blacklist", default=DEFAULT_BLACKLIST, help="키워드 블랙리스트 xlsx 경로")
    parser.add_argument("--tiers", default=DEFAULT_TIERS,
                        help=f"계단 구간 '상한:캡' 목록. 상한 0=무제한(마지막만). default: {DEFAULT_TIERS}")
    parser.add_argument("--relevance-cap", type=int, default=0,
                        help="상품명/대표키워드 어근과 겹치는 후보를 캡과 무관하게 최대 N개 우선 포함 "
                             "(0=비활성, 기존 동작). ⑤ product-name은 20 권장")
    parser.add_argument("--parent-relevance-cap", type=int, default=0,
                        help="상위 카테고리(2차 접두) 풀에서 상품 어근에 걸리는 후보를 최대 N개 추가 "
                             "(0=비활성). cat_prefilter --parent-fallback 로 만든 manifest가 필요하다")
    parser.add_argument("--parent-min-roots", type=int, default=2,
                        help="상위 후보가 포함해야 할 최소 상품 어근 수 (default: 2). "
                             "1로 낮추면 범용 어근 1개만 걸린 무관 상품이 대량 유입된다")
    args = parser.parse_args()

    try:
        tiers = parse_tiers(args.tiers)
    except ValueError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)

    run_dir = Path(args.run_dir)
    manifest_path = run_dir / "manifest.json"
    targets_path = run_dir / "targets.json"
    out_dir = run_dir / "candidates"

    if not manifest_path.exists():
        print(f"ERROR: manifest.json을 찾을 수 없습니다: {manifest_path}", file=sys.stderr)
        sys.exit(1)
    if not targets_path.exists():
        print(f"ERROR: targets.json을 찾을 수 없습니다: {targets_path}", file=sys.stderr)
        sys.exit(1)

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        targets = json.loads(targets_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        print(f"ERROR: JSON 파싱 실패: {e}", file=sys.stderr)
        sys.exit(1)

    categories = manifest.get("categories", {})
    parents_meta = manifest.get("parents", {})
    excluded_keywords, brand_roots = load_blacklist(args.blacklist)

    out_dir.mkdir(parents=True, exist_ok=True)

    # 카테고리별 후보는 한 번만 계산해 재사용
    category_cache = {}   # 카테고리 -> 계단 캡 적용 후보(공유 풀)
    category_pool = {}    # 카테고리 -> 계단 태깅된 전체 풀(상품별 관련어 선별용)
    category_link = {}
    parent_pool = {}      # p2norm -> 계단 태깅된 상위 풀
    products = []
    skipped = []
    rel_stat = []
    par_stat = []

    def get_parent_pool(p2norm):
        """상위 묶음 xlsx를 한 번만 로드해 계단 태깅된 풀로 캐시한다."""
        if p2norm in parent_pool:
            return parent_pool[p2norm]
        parent_pool[p2norm] = []
        info = parents_meta.get(p2norm)
        if not info:
            return parent_pool[p2norm]
        ppath = Path(info.get("file", ""))
        if not ppath.is_absolute():
            ppath = run_dir / ppath
        if not ppath.exists():
            print(f"경고: 상위 묶음 파일이 없어 건너뜁니다: {ppath}", file=sys.stderr)
            return parent_pool[p2norm]
        try:
            _, parent_pool[p2norm] = build_capped_candidates(
                str(ppath), excluded_keywords, brand_roots, tiers)
        except ValueError as e:
            print(f"경고: 상위 묶음 '{p2norm}' 후보 생성 실패({e}) — 건너뜁니다.", file=sys.stderr)
        return parent_pool[p2norm]

    for t in targets:
        product_id = t.get("productId")
        category = t.get("카테고리")
        cat_info = categories.get(category)
        if cat_info is None:
            print(f"경고: 상품 {product_id}의 카테고리 '{category}'가 manifest에 없어 건너뜁니다.", file=sys.stderr)
            skipped.append(product_id)
            continue

        if category not in category_cache:
            file_field = cat_info.get("file")
            xlsx_path = Path(file_field)
            if not xlsx_path.is_absolute():
                xlsx_path = run_dir / xlsx_path
            if not xlsx_path.exists():
                print(f"경고: 카테고리 '{category}' 파일을 찾을 수 없어 건너뜁니다: {xlsx_path}", file=sys.stderr)
                category_cache[category] = []
                category_pool[category] = []
                category_link[category] = str(xlsx_path)
                continue
            try:
                category_cache[category], category_pool[category] = build_capped_candidates(
                    str(xlsx_path), excluded_keywords, brand_roots, tiers)
            except ValueError as e:
                print(f"경고: 카테고리 '{category}' 후보 생성 실패({e}) — 건너뜁니다.", file=sys.stderr)
                category_cache[category] = []
                category_pool[category] = []
            # 원본파일링크: 실제 Drive 링크는 run_batch가 나중에 주입. 여기선 소스 파일 경로를 placeholder로 둔다.
            category_link[category] = str(xlsx_path)

        # 상품 관련어 우선 선별 — 계단 캡에서 잘려나간 직결 키워드를 살린다.
        # (저상품수 구간은 브랜드/모델명이 지배하므로 상품수 정렬만으로는 직결어가 체계적으로 배제됨)
        roots = extract_roots(t.get("상품명"), t.get("대표키워드"))
        relevant = pick_relevant(category_pool.get(category, []), roots, args.relevance_cap)
        rel_slim = [{k: c.get(k) for k in CANDIDATE_FIELDS} for c in relevant]
        for c in rel_slim:
            c["상품관련"] = True

        # 상위 카테고리 폴백 — leaf 풀에 그 상품 직결어가 거의 없는 경우를 메운다.
        # (leaf 후보를 대체하지 않고 뒤에 덧붙인다. `기준`에 '상위카테고리' 표기로 구분)
        par_slim = []
        p2norm = cat_info.get("parent")
        if args.parent_relevance_cap and p2norm:
            leaf_kws = ({c["키워드"] for c in rel_slim}
                        | {c["키워드"] for c in category_cache[category]})
            par = pick_parent_relevant(get_parent_pool(p2norm), roots,
                                       args.parent_relevance_cap, category,
                                       exclude=leaf_kws, min_roots=args.parent_min_roots)
            par_slim = [{k: c.get(k) for k in CANDIDATE_FIELDS} for c in par]
            for c in par_slim:
                c["상품관련"] = True
        par_stat.append(len(par_slim))

        merged, seen = [], set()
        for c in rel_slim + par_slim + list(category_cache[category]):
            if c["키워드"] in seen:
                continue
            seen.add(c["키워드"])
            merged.append(c)
        rel_stat.append(len(rel_slim))

        products.append({
            "productId": product_id,
            "상품명": t.get("상품명"),
            "대표키워드": t.get("대표키워드"),
            "카테고리": category,
            "원본파일링크": category_link[category],
            "candidates": merged,
        })

    # 상품 10개/청크로 분할 (상품 경계로만 자름)
    chunk_size = args.chunk_size
    chunk_files = []
    for i in range(0, len(products), chunk_size):
        chunk_no = i // chunk_size + 1
        chunk = {"chunk": chunk_no, "products": products[i:i + chunk_size]}
        chunk_path = out_dir / f"chunk_{chunk_no:03d}.json"
        chunk_path.write_text(json.dumps(chunk, ensure_ascii=False, indent=2), encoding="utf-8")
        chunk_files.append(chunk_path)

    print(f"# 카테고리: {len(category_cache)}개 처리")
    if args.relevance_cap and rel_stat:
        zero = sum(1 for n in rel_stat if n == 0)
        print(f"# 상품관련 후보: 평균 {sum(rel_stat)/len(rel_stat):.1f}개/상품, "
              f"0건인 상품 {zero}개 (관련어 매칭 실패 → 보류 후보)")
    if args.parent_relevance_cap and par_stat:
        pzero = sum(1 for n in par_stat if n == 0)
        print(f"# 상위카테고리 후보: 평균 {sum(par_stat)/len(par_stat):.1f}개/상품, "
              f"0건인 상품 {pzero}개 (상위 묶음 {len(parents_meta)}개)")
    print(f"# 상품: {len(products)}개 청크화 (건너뜀 {len(skipped)}개)")
    print(f"# 청크: {len(chunk_files)}개 → {out_dir}")
    for p in chunk_files:
        print(f"  - {p.name}")


if __name__ == "__main__":
    main()
