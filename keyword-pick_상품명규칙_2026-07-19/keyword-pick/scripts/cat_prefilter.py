#!/usr/bin/env python3
"""셀러라이프 통다운(전체 카테고리 ZIP/raw)에서 여러 상품의 서로 다른 카테고리 행을
파일당 1회 스캔으로 동시 수집한다.

기존 프로토타입(scratchpad/cat_prefilter.py)은 "상품(카테고리) 1개당 파일 1스캔"이라
846개 상품이면 비효율적이었다. 이 스크립트는 "파일당 1스캔, 출력은 카테고리별"로
동작해 필요한 파일만 최대 1회씩 훑는다.

입력: targets.json ([{"productId","상품명","대표키워드","카테고리"}, ...])
출력:
  <out-dir>/prefiltered/<카테고리슬러그>.xlsx  (카테고리당 1파일, 원본 35열 헤더 보존, 시트명 all)
  <out-dir>/manifest.json                     (카테고리→파일/행수/productIds, not_found, 캐시 키)

사용법:
  python cat_prefilter.py --targets targets.json --out-dir <dir> --raw-dir <dir> [--zip <zip경로>]
                           [--precut 100000] [--force]

캐시: manifest.json에 (통다운 source_date × targets 해시)를 기록해두고, 동일 조합으로
재실행되면 스캔을 건너뛴다. --force로 강제 재스캔 가능.
"""
import argparse
import hashlib
import json
import os
import re
import sys
import zipfile
from datetime import date

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
from kwlib import sanitize_part  # noqa: E402

# common.py(cat_key)는 sellerlife-keyword 스킬 소유 — bulsaja가 sellha.py를 빌려쓰는 관례와 동일하게 sys.path로 차용
SELLERLIFE_SCRIPTS = os.path.join(
    os.path.dirname(os.path.dirname(SCRIPT_DIR)), "sellerlife-keyword", "scripts"
)
sys.path.insert(0, SELLERLIFE_SCRIPTS)
from common import cat_key  # noqa: E402

try:
    import openpyxl
    from openpyxl import Workbook
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl>=3.1.0", file=sys.stderr)
    sys.exit(1)


# 대분류(카테고리 1차 세그먼트) → 통다운 파일 stem(SellarSourcing_<stem>.xlsx) 매핑.
# '생활/건강'은 20만행 export 상한으로 2개 파일에 나뉘어 있어 둘 다 스캔한다.
# 여기 없는 대분류(면세점 등)는 통다운 커버 밖으로 판정한다.
MAJOR_TO_STEMS = {
    "가구/인테리어": ["가구&인테리어"],
    "디지털/가전": ["디지털&가전"],
    "생활/건강": ["생활&건강", "생활&건강2"],
    "도서": ["도서"],
    "식품": ["식품"],
    "스포츠/레저": ["스포츠&레저"],
    "여가/생활편의": ["여가&생활편의"],
    "출산/육아": ["출산&육아"],
    "패션의류": ["패션의류"],
    "패션잡화": ["패션잡화"],
    "화장품/미용": ["화장품&미용"],
}
MAJOR_KEY_TO_STEMS = {cat_key(k): v for k, v in MAJOR_TO_STEMS.items()}


def nows(s):
    """공백·줄바꿈 제거 정규화 (카테고리 정확일치 비교용)."""
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def slugify_category(cat_str):
    """카테고리 경로를 파일 안전 슬러그로. '디지털/가전>주방가전>빙수기' -> '디지털가전_주방가전_빙수기'."""
    parts = [sanitize_part(p) for p in str(cat_str).split(">")]
    parts = [p for p in parts if p]
    return "_".join(parts) if parts else "미분류"


def _decode_member(name, flag_bits):
    """ZIP 멤버명 한글 깨짐 복구. UTF-8 플래그가 서 있으면 그대로, 아니면 cp437->cp949 복구 시도.
    (sellerlife-keyword/ingest.py의 _decode_member와 동일 로직 차용)"""
    if flag_bits & 0x800:
        return name
    try:
        return name.encode("cp437").decode("cp949")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return name


def detect_source_date(raw_dir, zip_path):
    """통다운 날짜(YYMMDD)를 raw_dir 경로(runs/<YYMMDD>/raw) 또는 zip 파일명(<YYMMDD>_...) 에서 추출."""
    if raw_dir:
        m = re.search(r"runs[\\/](\d{6})[\\/]", raw_dir.replace("/", os.sep) + os.sep)
        if m:
            return m.group(1)
    if zip_path:
        m = re.match(r"^(\d{6})_", os.path.basename(zip_path))
        if m:
            return m.group(1)
    return date.today().strftime("%y%m%d")


def compute_targets_hash(targets):
    """targets(productId, 카테고리) 조합의 해시. 캐시 히트 판정용."""
    pairs = sorted(
        f"{t.get('productId', '')}|{t.get('카테고리', '')}" for t in targets
    )
    h = hashlib.sha256("\n".join(pairs).encode("utf-8"))
    return h.hexdigest()[:16]


def resolve_stem_file(stem, raw_dir, zip_path):
    """stem(예: '디지털&가전')에 해당하는 실제 파일 경로를 확보한다.
    raw_dir에 이미 있으면 재사용, 없으면 zip에서 추출한다."""
    basename = f"SellarSourcing_{stem}.xlsx"
    dest = os.path.join(raw_dir, basename)
    if os.path.exists(dest):
        return dest

    if not zip_path or not os.path.exists(zip_path):
        raise RuntimeError(f"'{basename}' 이 raw-dir에 없고, zip 경로도 없어 확보할 수 없습니다.")

    with zipfile.ZipFile(zip_path) as z:
        for info in z.infolist():
            if info.is_dir():
                continue
            name = _decode_member(info.filename, info.flag_bits)
            if os.path.basename(name) == basename:
                os.makedirs(raw_dir, exist_ok=True)
                with z.open(info) as src, open(dest, "wb") as out:
                    out.write(src.read())
                return dest
    raise RuntimeError(f"zip 안에서 '{basename}' 을 찾지 못했습니다.")


def route_targets(targets):
    """targets를 대분류 기준으로 라우팅.
    반환: (routed_categories: {카테고리원문: {"norm","productIds","stems"}}, not_found: [...])
    """
    routed = {}  # norm -> {"repr":원문, "productIds":[...], "stems":set}
    not_found = []

    for t in targets:
        pid = t.get("productId", "")
        cat = str(t.get("카테고리", "") or "").strip()
        if not cat:
            not_found.append({"productId": pid, "카테고리": cat, "사유": "카테고리없음"})
            continue

        major = cat.split(">")[0].strip()
        mkey = cat_key(major)
        stems = MAJOR_KEY_TO_STEMS.get(mkey)
        if not stems:
            not_found.append({"productId": pid, "카테고리": cat, "사유": "통다운밖"})
            continue

        norm = nows(cat)
        if norm not in routed:
            routed[norm] = {"repr": cat, "productIds": [], "stems": set()}
        routed[norm]["productIds"].append(pid)
        routed[norm]["stems"].update(stems)

    return routed, not_found


def scan_and_collect(needed_files, routed, precut):
    """필요한 파일들을 각 1회씩 스트리밍 스캔해 카테고리별 버킷(35열 원본행)에 담는다.
    반환: (buckets: {norm: [row,...]}, prefix_counts: {norm: int})
    """
    exact_index = set(routed.keys())
    # 3차 접두 진단용: norm -> prefix3norm
    prefix3_of = {}
    prefix_index = {}  # prefix3norm -> [norm, ...]
    for norm, info in routed.items():
        parts = info["repr"].split(">")
        p3 = ">".join(parts[:3]) if len(parts) >= 3 else info["repr"]
        p3norm = nows(p3)
        prefix3_of[norm] = p3norm
        prefix_index.setdefault(p3norm, []).append(norm)

    buckets = {}
    prefix_counts = {norm: 0 for norm in routed}
    header = None
    cat_i = None

    for pth in needed_files:
        wb = openpyxl.load_workbook(pth, read_only=True, data_only=True)
        try:
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            file_header = next(it)
            if header is None:
                header = list(file_header)
                for i, h in enumerate(header):
                    if "카테고리" in nows(h):
                        cat_i = i
                        break
            for row in it:
                if cat_i is None or cat_i >= len(row):
                    continue
                rnorm = nows(row[cat_i])
                if rnorm in exact_index:
                    bucket = buckets.setdefault(rnorm, [])
                    if precut <= 0 or len(bucket) < precut:
                        bucket.append(list(row))
                else:
                    rparts = rnorm.split(">")
                    rp3 = ">".join(rparts[:3])
                    for norm in prefix_index.get(rp3, ()):
                        prefix_counts[norm] += 1
        finally:
            wb.close()

    return header, buckets, prefix_counts


def write_prefiltered(out_dir, header, routed, buckets):
    """카테고리별 버킷을 <out-dir>/prefiltered/<슬러그>.xlsx 로 저장. categories manifest 항목 생성."""
    pref_dir = os.path.join(out_dir, "prefiltered")
    os.makedirs(pref_dir, exist_ok=True)

    categories = {}
    for norm, info in routed.items():
        rows = buckets.get(norm)
        if not rows:
            continue  # 0매치는 여기서 스킵 — not_found(표기불일치의심)에서 처리
        slug = slugify_category(info["repr"])
        fname = f"{slug}.xlsx"
        fpath = os.path.join(pref_dir, fname)
        wb_out = Workbook()
        ws_out = wb_out.active
        ws_out.title = "all"
        ws_out.append(header)
        for row in rows:
            ws_out.append(row)
        wb_out.save(fpath)

        categories[info["repr"]] = {
            "file": f"prefiltered/{fname}",
            "rows": len(rows),
            "productIds": info["productIds"],
        }
    return categories


def main():
    ap = argparse.ArgumentParser(description="셀러라이프 통다운 다중 카테고리 프리필터")
    ap.add_argument("--targets", required=True, help="targets.json 경로")
    ap.add_argument("--out-dir", required=True, help="prefiltered/·manifest.json 저장 폴더")
    ap.add_argument("--raw-dir", required=True, help="통다운 raw xlsx 폴더(이미 풀린 파일 재사용/추출 대상)")
    ap.add_argument("--zip", default=None, help="통다운 전체카테고리 zip 경로 (미지정시 raw-dir에서 *.zip 자동탐색)")
    ap.add_argument("--precut", type=int, default=100000, help="카테고리당 최대 수집 행수(메모리 방어). 0=무컷")
    ap.add_argument("--force", action="store_true", help="캐시 무시하고 강제 재스캔")
    args = ap.parse_args()

    with open(args.targets, encoding="utf-8") as f:
        targets = json.load(f)
    if not targets:
        print("targets.json이 비어있습니다.")
        sys.exit(1)

    zip_path = args.zip
    if not zip_path:
        for fn in os.listdir(args.raw_dir):
            if fn.lower().endswith(".zip"):
                zip_path = os.path.join(args.raw_dir, fn)
                break

    source_date = detect_source_date(args.raw_dir, zip_path)
    targets_hash = compute_targets_hash(targets)

    os.makedirs(args.out_dir, exist_ok=True)
    manifest_path = os.path.join(args.out_dir, "manifest.json")

    # 캐시 체크: (source_date, targets_hash) 동일하면 재스캔 스킵
    if not args.force and os.path.exists(manifest_path):
        try:
            with open(manifest_path, encoding="utf-8") as f:
                cached = json.load(f)
            if cached.get("source_date") == source_date and cached.get("targets_hash") == targets_hash:
                print(f"###CACHE_HIT### source_date={source_date} targets_hash={targets_hash} -> 재스캔 스킵")
                print(f"manifest: {manifest_path}")
                return
        except (json.JSONDecodeError, OSError):
            pass  # 캐시 파일 깨졌으면 그냥 재스캔

    # 1) 라우팅
    routed, not_found = route_targets(targets)

    # 2) 필요한 파일 목록 확보 (중복 제거, stem 단위)
    needed_stems = set()
    for info in routed.values():
        needed_stems.update(info["stems"])

    needed_files = []
    for stem in sorted(needed_stems):
        try:
            fpath = resolve_stem_file(stem, args.raw_dir, zip_path)
            needed_files.append(fpath)
            print(f"[파일확보] {stem} -> {fpath}")
        except RuntimeError as e:
            print(f"[오류] {e}")
            # 이 stem에 걸린 카테고리들은 스캔 자체가 불가능 -> not_found로 이동
            for norm, info in list(routed.items()):
                if stem in info["stems"]:
                    for pid in info["productIds"]:
                        not_found.append({"productId": pid, "카테고리": info["repr"], "사유": "파일확보실패"})
                    del routed[norm]

    # 3) 스캔 (파일당 1회)
    header, buckets, prefix_counts = scan_and_collect(needed_files, routed, args.precut)

    # 4) 저장
    categories = write_prefiltered(args.out_dir, header, routed, buckets) if header else {}

    # 5) 0매치 카테고리는 표기불일치의심으로 not_found에 추가
    for norm, info in routed.items():
        if not buckets.get(norm):
            for pid in info["productIds"]:
                not_found.append({
                    "productId": pid,
                    "카테고리": info["repr"],
                    "사유": "표기불일치의심",
                    "참고": f"3차접두 매칭(참고)={prefix_counts.get(norm, 0)}건",
                })

    manifest = {
        "categories": categories,
        "not_found": not_found,
        "source_date": source_date,
        "targets_hash": targets_hash,
    }
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    total_rows = sum(v["rows"] for v in categories.values())
    print(f"###DONE### 카테고리 {len(categories)}건 저장, 총 {total_rows}행 / not_found {len(not_found)}건")
    print(f"manifest: {manifest_path}")


if __name__ == "__main__":
    main()
