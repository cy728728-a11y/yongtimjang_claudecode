#!/usr/bin/env python3
"""프리필터 결과(카테고리별 xlsx) → 상품별 계단 태깅 후보 JSON 청크 생성.

키워드→상품명 배치 파이프라인 3단계 스크립트.
kwlib.build_candidates()로 후보를 뽑고, 상품수 구간(1만/2만/3만)으로 계단 태깅한 뒤
캡을 적용해 상품당 최대 35개(15+10+10)로 줄인다. 여기서 브랜드/직결여부 판정은 하지 않는다
(그건 Claude가 chunk를 읽고 판단). 이 스크립트는 기계적 준비만 한다.

입력 (run_dir 안에):
  - manifest.json : {"categories": {"<카테고리>": {"file": "prefiltered/xxx.xlsx"}}}
    file 경로는 run_dir 기준 상대경로 또는 절대경로 모두 허용.
  - targets.json  : [{"productId":"...", "상품명":"...", "대표키워드":"...", "카테고리":"..."}, ...]

출력:
  - <run_dir>/candidates/chunk_001.json, chunk_002.json, ... (상품 10개/청크, 경계로만 자름)

Usage:
  python batch_candidates.py <run_dir> [--chunk-size 10] [--max-products 30000] [--blacklist PATH]
"""

import argparse
import json
import os
import sys
from pathlib import Path

# kwlib.py가 같은 폴더에 있으므로 어느 경로에서 실행해도 import 가능하도록 경로 추가
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from kwlib import build_candidates, load_rows

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl>=3.1.0", file=sys.stderr)
    sys.exit(1)

DEFAULT_BLACKLIST = r"D:\python_work\data\sellerlife\keyword_blacklist\keyword_blacklist.xlsx"

# 계단 구간 상한(상품수) 및 구간별 캡
TIER1_MAX = 10000
TIER2_MAX = 20000
TIER3_MAX = 30000
TIER1_CAP = 15
TIER2_CAP = 10
TIER3_CAP = 10

# 최종 출력에 남길 후보 필드(스키마 고정)
CANDIDATE_FIELDS = [
    "키워드", "총검색수", "상품수", "기준",
    "네이버해외배송비율", "쿠팡해외배송비율", "신규진입키워드", "브랜드의심",
]


def load_blacklist(path):
    """sellerlife 키워드 블랙리스트에서 제외키워드(정확일치)·제외브랜드(부분일치)를 읽는다.
    파일/시트가 없으면 조용히 빈 값 반환(설계문서 규칙)."""
    excluded_keywords = set()
    brand_roots = []
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return excluded_keywords, brand_roots

    try:
        if "제외키워드" in wb.sheetnames:
            ws = wb["제외키워드"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(h).strip() if h is not None else "" for h in rows[0]]
                kw_idx = header.index("키워드") if "키워드" in header else 0
                for r in rows[1:]:
                    if kw_idx < len(r) and r[kw_idx] is not None and str(r[kw_idx]).strip():
                        excluded_keywords.add(str(r[kw_idx]).strip())

        if "제외브랜드" in wb.sheetnames:
            ws = wb["제외브랜드"]
            rows = list(ws.iter_rows(values_only=True))
            if rows:
                header = [str(h).strip() if h is not None else "" for h in rows[0]]
                br_idx = header.index("브랜드어근") if "브랜드어근" in header else 0
                for r in rows[1:]:
                    if br_idx < len(r) and r[br_idx] is not None and str(r[br_idx]).strip():
                        brand_roots.append(str(r[br_idx]).strip())
    except Exception:
        # 시트 구조가 예상과 달라도 배치 전체를 막지 않는다 — 플래그 없이 진행
        pass
    finally:
        wb.close()

    return excluded_keywords, brand_roots


def is_brand_suspect(keyword, excluded_keywords, brand_roots):
    """블랙리스트 힌트 판정. 정확일치(제외키워드) 또는 부분일치(브랜드어근)면 True."""
    if keyword in excluded_keywords:
        return True
    for root in brand_roots:
        if root and root in keyword:
            return True
    return False


def build_capped_candidates(xlsx_path, excluded_keywords, brand_roots):
    """카테고리 xlsx 하나에서 계단 태깅 + 캡 적용된 후보 목록을 만든다."""
    idx, data = load_rows(xlsx_path)
    all_candidates = build_candidates(idx, data, TIER3_MAX, exclude_info=True)

    for c in all_candidates:
        pc = c["상품수"]
        if pc <= TIER1_MAX:
            c["기준"] = "상품수 1만 이하"
        elif pc <= TIER2_MAX:
            c["기준"] = "상품수 2만 이하 확장"
        else:
            c["기준"] = "상품수 3만 이하 확장"
        c["브랜드의심"] = is_brand_suspect(c["키워드"], excluded_keywords, brand_roots)

    all_candidates.sort(key=lambda c: c["상품수"])

    tier1 = [c for c in all_candidates if c["기준"] == "상품수 1만 이하"][:TIER1_CAP]
    tier2 = [c for c in all_candidates if c["기준"] == "상품수 2만 이하 확장"][:TIER2_CAP]
    tier3 = [c for c in all_candidates if c["기준"] == "상품수 3만 이하 확장"][:TIER3_CAP]
    capped = tier1 + tier2 + tier3

    # 스키마 고정 필드만 남긴다 (청크 용량 절약)
    slim = [{k: c.get(k) for k in CANDIDATE_FIELDS} for c in capped]
    return slim


def main():
    parser = argparse.ArgumentParser(description="프리필터 결과 → 상품별 계단태깅 후보 JSON 청크 생성")
    parser.add_argument("run_dir", help="manifest.json/targets.json이 있는 실행 디렉토리. candidates/ 하위에 출력")
    parser.add_argument("--chunk-size", type=int, default=10, help="청크당 상품 수 (default: 10)")
    parser.add_argument("--blacklist", default=DEFAULT_BLACKLIST, help="키워드 블랙리스트 xlsx 경로")
    args = parser.parse_args()

    run_dir = Path(args.run_dir)
    manifest_path = run_dir / "manifest.json"
    targets_path = run_dir / "targets.json"
    out_dir = run_dir / "candidates"

    if not manifest_path.exists():
        print(f"ERROR: manifest.json을 찾을 수 없습니다: {manifest_path}", file=sys.stderr)
        sys.exit(1)
    if not targets_path.exists():
        print(f"ERROR: targets.json을 찾을 수 없습니다: {targets_path}", file=sys.stderr)
        sys.exit(1)

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
        targets = json.loads(targets_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError as e:
        print(f"ERROR: JSON 파싱 실패: {e}", file=sys.stderr)
        sys.exit(1)

    categories = manifest.get("categories", {})
    excluded_keywords, brand_roots = load_blacklist(args.blacklist)

    out_dir.mkdir(parents=True, exist_ok=True)

    # 카테고리별 후보는 한 번만 계산해 재사용
    category_cache = {}
    category_link = {}
    products = []
    skipped = []

    for t in targets:
        product_id = t.get("productId")
        category = t.get("카테고리")
        cat_info = categories.get(category)
        if cat_info is None:
            print(f"경고: 상품 {product_id}의 카테고리 '{category}'가 manifest에 없어 건너뜁니다.", file=sys.stderr)
            skipped.append(product_id)
            continue

        if category not in category_cache:
            file_field = cat_info.get("file")
            xlsx_path = Path(file_field)
            if not xlsx_path.is_absolute():
                xlsx_path = run_dir / xlsx_path
            if not xlsx_path.exists():
                print(f"경고: 카테고리 '{category}' 파일을 찾을 수 없어 건너뜁니다: {xlsx_path}", file=sys.stderr)
                category_cache[category] = []
                category_link[category] = str(xlsx_path)
                continue
            try:
                category_cache[category] = build_capped_candidates(str(xlsx_path), excluded_keywords, brand_roots)
            except ValueError as e:
                print(f"경고: 카테고리 '{category}' 후보 생성 실패({e}) — 건너뜁니다.", file=sys.stderr)
                category_cache[category] = []
            # 원본파일링크: 실제 Drive 링크는 run_batch가 나중에 주입. 여기선 소스 파일 경로를 placeholder로 둔다.
            category_link[category] = str(xlsx_path)

        products.append({
            "productId": product_id,
            "상품명": t.get("상품명"),
            "대표키워드": t.get("대표키워드"),
            "카테고리": category,
            "원본파일링크": category_link[category],
            "candidates": list(category_cache[category]),  # 얕은 복사(카테고리별 공유 풀, 상품마다 독립 리스트)
        })

    # 상품 10개/청크로 분할 (상품 경계로만 자름)
    chunk_size = args.chunk_size
    chunk_files = []
    for i in range(0, len(products), chunk_size):
        chunk_no = i // chunk_size + 1
        chunk = {"chunk": chunk_no, "products": products[i:i + chunk_size]}
        chunk_path = out_dir / f"chunk_{chunk_no:03d}.json"
        chunk_path.write_text(json.dumps(chunk, ensure_ascii=False, indent=2), encoding="utf-8")
        chunk_files.append(chunk_path)

    print(f"# 카테고리: {len(category_cache)}개 처리")
    print(f"# 상품: {len(products)}개 청크화 (건너뜀 {len(skipped)}개)")
    print(f"# 청크: {len(chunk_files)}개 → {out_dir}")
    for p in chunk_files:
        print(f"  - {p.name}")


if __name__ == "__main__":
    main()
